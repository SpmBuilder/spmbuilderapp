import streamlit as st
import pandas as pd
import numpy as np
import json
import re
from pathlib import Path
import io
import base64
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy as _copy_style
import zipfile
import tempfile
import os

# Try to import msoffcrypto for encrypted Excel files
try:
    import msoffcrypto
    HAS_MSOFFCRYPTO = True
except ImportError:
    HAS_MSOFFCRYPTO = False

# Page configuration
st.set_page_config(
    page_title="Multi-Sheet Data Processing Wizard",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap');

    .stApp {
        background: radial-gradient(circle at 50% -20%, #1c233a 0%, #0d111d 45%, #06080e 100%) !important;
        color: #e2e8f0 !important;
        font-family: 'Inter', sans-serif !important;
    }

    [data-testid="stSidebar"] {
        background-color: #0c0f17 !important;
        border-right: 1px solid rgba(255, 255, 255, 0.05) !important;
    }
    [data-testid="stSidebar"] * {
        font-family: 'Outfit', sans-serif !important;
    }

    .main-header {
        font-family: 'Outfit', sans-serif !important;
        font-size: 2.8rem !important;
        font-weight: 800 !important;
        background: linear-gradient(135deg, #00f2fe 0%, #4facfe 50%, #9b51e0 100%) !important;
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
        text-align: center !important;
        margin-bottom: 0.2rem !important;
        letter-spacing: -0.5px !important;
        text-shadow: 0 0 40px rgba(0, 242, 254, 0.15);
    }

    .main-subheader {
        font-family: 'Outfit', sans-serif !important;
        font-size: 1.1rem !important;
        color: #8a99ad !important;
        text-align: center !important;
        margin-bottom: 2rem !important;
        font-weight: 300 !important;
    }

    .glass-card {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.03) 0%, rgba(255, 255, 255, 0.005) 100%) !important;
        backdrop-filter: blur(16px) !important;
        -webkit-backdrop-filter: blur(16px) !important;
        border: 1px solid rgba(255, 255, 255, 0.05) !important;
        border-top: 1px solid rgba(0, 242, 254, 0.15) !important;
        border-left: 1px solid rgba(0, 242, 254, 0.08) !important;
        border-radius: 12px !important;
        padding: 24px !important;
        margin-bottom: 20px !important;
        box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.4) !important;
    }

    .password-box {
        background: linear-gradient(135deg, rgba(255, 165, 0, 0.08) 0%, rgba(255, 165, 0, 0.02) 100%) !important;
        border: 1px solid rgba(255, 165, 0, 0.3) !important;
        border-top: 1px solid rgba(255, 165, 0, 0.5) !important;
        border-radius: 12px !important;
        padding: 20px !important;
        margin: 15px 0 !important;
    }

    .password-box h4 {
        color: #ffa500 !important;
        font-family: 'Outfit', sans-serif !important;
    }

    .date-shortcut-box {
        background: linear-gradient(135deg, rgba(0, 242, 254, 0.05) 0%, rgba(0, 242, 254, 0.01) 100%) !important;
        border: 1px solid rgba(0, 242, 254, 0.15) !important;
        border-radius: 8px !important;
        padding: 12px !important;
        margin: 8px 0 !important;
    }

    .header-review-box {
        background: linear-gradient(135deg, rgba(155, 81, 224, 0.08) 0%, rgba(155, 81, 224, 0.02) 100%) !important;
        border: 1px solid rgba(155, 81, 224, 0.25) !important;
        border-radius: 12px !important;
        padding: 16px !important;
        margin: 12px 0 !important;
    }

    .header-review-box h4 {
        color: #9b51e0 !important;
        font-family: 'Outfit', sans-serif !important;
    }

    .welcome-box {
        background: linear-gradient(135deg, rgba(0, 242, 254, 0.08) 0%, rgba(79, 172, 254, 0.02) 50%, rgba(155, 81, 224, 0.08) 100%) !important;
        border: 1px solid rgba(0, 242, 254, 0.15) !important;
        border-top: 1px solid rgba(0, 242, 254, 0.35) !important;
        border-radius: 16px !important;
        padding: 35px !important;
        margin: 20px 0 !important;
        box-shadow: 0 10px 40px rgba(0, 0, 0, 0.45) !important;
    }

    .operation-active {
        background: linear-gradient(135deg, rgba(155, 81, 224, 0.08) 0%, rgba(225, 0, 255, 0.02) 100%) !important;
        backdrop-filter: blur(16px) !important;
        border: 1px solid rgba(155, 81, 224, 0.3) !important;
        border-top: 1px solid rgba(225, 0, 255, 0.5) !important;
        border-radius: 12px !important;
        padding: 22px !important;
        margin: 20px 0 !important;
        box-shadow: 0 0 25px rgba(155, 81, 224, 0.2) !important;
    }

    [data-testid="stMetric"] {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.03) 0%, rgba(255, 255, 255, 0.005) 100%) !important;
        border: 1px solid rgba(255, 255, 255, 0.04) !important;
        border-top: 1px solid rgba(0, 242, 254, 0.15) !important;
        border-radius: 10px !important;
        padding: 12px 18px !important;
        text-align: center !important;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2) !important;
    }

    div.stButton > button {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.03) 0%, rgba(255, 255, 255, 0.005) 100%) !important;
        color: #cbd5e1 !important;
        border: 1px solid rgba(255, 255, 255, 0.08) !important;
        border-radius: 8px !important;
        padding: 8px 16px !important;
        font-weight: 500 !important;
        font-family: 'Inter', sans-serif !important;
        transition: all 0.25s ease !important;
        width: 100% !important;
    }

    div.stButton > button:hover {
        background: linear-gradient(135deg, rgba(0, 242, 254, 0.1) 0%, rgba(155, 81, 224, 0.1) 100%) !important;
        border-color: rgba(0, 242, 254, 0.4) !important;
        color: #ffffff !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 15px rgba(0, 242, 254, 0.15) !important;
    }

    div.stButton > button[key*="apply_"], 
    div.stButton > button[key*="run_"], 
    div.stButton > button[key*="tpl_build_"], 
    div.stButton > button[key*="load_all_files_btn"],
    div.stButton > button[key*="download_"] {
        background: linear-gradient(135deg, #9b51e0 0%, #e100ff 50%, #00f2fe 100%) !important;
        color: #ffffff !important;
        border: none !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 15px rgba(225, 0, 255, 0.25) !important;
    }

    .navbar-container {
        background: rgba(255, 255, 255, 0.01) !important;
        border: 1px solid rgba(255, 255, 255, 0.05) !important;
        border-radius: 14px !important;
        padding: 10px !important;
        margin-bottom: 25px !important;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.15) !important;
    }

    .timeline-step {
        border-left: 2px solid #9b51e0 !important;
        padding-left: 18px !important;
        margin-left: 10px !important;
        margin-bottom: 20px !important;
        position: relative !important;
    }
    .timeline-step::before {
        content: '●' !important;
        position: absolute !important;
        left: -7px !important;
        top: -1px !important;
        color: #00f2fe !important;
        font-size: 14px !important;
        text-shadow: 0 0 8px rgba(0, 242, 254, 0.8) !important;
    }
    .timeline-step-content {
        background: rgba(255, 255, 255, 0.01) !important;
        border: 1px solid rgba(255, 255, 255, 0.03) !important;
        border-radius: 8px !important;
        padding: 10px 14px !important;
    }

    .manual-grid {
        display: grid !important;
        grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)) !important;
        gap: 20px !important;
        margin: 25px 0 !important;
    }

    .manual-card {
        background: rgba(255, 255, 255, 0.02) !important;
        backdrop-filter: blur(8px) !important;
        -webkit-backdrop-filter: blur(8px) !important;
        border: 1px solid rgba(255, 255, 255, 0.05) !important;
        border-radius: 12px !important;
        padding: 24px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2) !important;
        display: flex !important;
        flex-direction: column !important;
        height: 100% !important;
    }

    .manual-card:hover {
        transform: translateY(-2px) !important;
        border-color: rgba(0, 242, 254, 0.2) !important;
        box-shadow: 0 8px 30px rgba(0, 242, 254, 0.1) !important;
        background: rgba(255, 255, 255, 0.03) !important;
    }

    .manual-card-header {
        font-family: 'Outfit', sans-serif !important;
        font-size: 1.15rem !important;
        font-weight: 700 !important;
        color: #00f2fe !important;
        margin-bottom: 12px !important;
        display: flex !important;
        align-items: center !important;
        gap: 10px !important;
    }

    .manual-card-header.step {
        color: #9b51e0 !important;
    }

    .manual-card-content {
        font-size: 0.9rem !important;
        color: #cbd5e1 !important;
        line-height: 1.6 !important;
        flex-grow: 1 !important;
    }

    .manual-badge {
        background: rgba(155, 81, 224, 0.1) !important;
        border: 1px solid rgba(155, 81, 224, 0.2) !important;
        color: #b57cff !important;
        padding: 2px 8px !important;
        border-radius: 6px !important;
        font-size: 0.75rem !important;
        font-weight: 600 !important;
    }
    </style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Session state initialization
# ---------------------------------------------------------------------------
defaults = {
    'current_page': 'Manual',
    'data_sources': {},
    'source_metadata': {},
    'operations': [],
    'active_source': None,
    'relationships': [],
    'current_operation': None,
    'loaded_recipe': None,
    'recipe_mapping': {},
    'template_headers': [],
    'template_source_file': None,
    'source_snapshots': {},
    'operations_undo_stack': [],
    'batch_results': {},
    'filter_conditions': [],
    'filter_logic': 'AND',
    'pending_header_review': [],
    'original_files': {},
    'pending_password_files': [],
    'password_cache': {},
    'show_nav_tip': True,
}
for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# ---------------------------------------------------------------------------
# Header-row detection and multi-sheet loading helpers
# ---------------------------------------------------------------------------
def is_likely_header_row(row, threshold=0.4):
    """Check if a row looks like a header row (mostly text, few numbers)."""
    if row is None or len(row) == 0:
        return False
    
    non_null = row.dropna()
    if len(non_null) == 0:
        return False
    
    # Count text vs numeric values
    text_count = 0
    numeric_count = 0
    for v in non_null:
        try:
            float(str(v).replace(',', '').replace('$', '').replace('%', ''))
            numeric_count += 1
        except (ValueError, TypeError):
            text_count += 1
    
    # Header rows typically have more text than numbers
    text_ratio = text_count / len(non_null) if len(non_null) > 0 else 0
    return text_ratio >= threshold


def detect_header_row(raw_preview, max_rows=20):
    """Detect the most likely header row from a raw preview."""
    if raw_preview is None or len(raw_preview) == 0:
        return 0
    
    # Limit to max_rows
    preview = raw_preview.head(min(max_rows, len(raw_preview)))
    
    best_score = -1
    best_row = 0
    
    for i in range(len(preview)):
        row = preview.iloc[i]
        # Check if row has enough non-null values
        non_null_count = row.count()
        if non_null_count < 2:  # Skip mostly empty rows
            continue
        
        # Score: how header-like is this row
        score = 0
        
        # Prefer rows with more non-null values
        score += non_null_count * 0.5
        
        # Prefer rows with more text (headers)
        text_count = 0
        for v in row.dropna():
            try:
                float(str(v).replace(',', '').replace('$', '').replace('%', ''))
            except (ValueError, TypeError):
                text_count += 1
        text_ratio = text_count / non_null_count if non_null_count > 0 else 0
        score += text_ratio * 5
        
        # Prefer rows with unique values (headers should be distinct)
        unique_ratio = row.nunique() / non_null_count if non_null_count > 0 else 0
        score += unique_ratio * 3
        
        if score > best_score:
            best_score = score
            best_row = i
    
    return best_row


def read_raw_preview_with_header(file_bytes, ext, sheet_name=None, nrows=20):
    """Read raw preview with header detection info."""
    try:
        bio = io.BytesIO(file_bytes)
        if ext == ".csv":
            df = pd.read_csv(bio, header=None, nrows=nrows, dtype=str)
            return df, "csv"
        else:
            # Try to read with openpyxl first to get sheet info
            xl = pd.ExcelFile(bio)
            if sheet_name is None:
                sheet_name = xl.sheet_names[0]
            
            # Read raw data
            bio.seek(0)
            df = pd.read_excel(bio, sheet_name=sheet_name, header=None, nrows=nrows, dtype=str)
            return df, sheet_name
    except Exception as e:
        return None, None


def load_sheet_with_header(file_bytes, ext, sheet_name, header_row):
    """Load a sheet with a specified header row."""
    try:
        bio = io.BytesIO(file_bytes)
        if ext == ".csv":
            df = pd.read_csv(bio, header=header_row)
        else:
            df = pd.read_excel(bio, sheet_name=sheet_name, header=header_row)
        
        df.columns = dedupe_column_names(df.columns)
        for col in df.columns:
            converted = try_convert_datetime_column(df[col])
            if converted is not None:
                df[col] = converted
        return df
    except Exception as e:
        st.error(f"Error loading sheet with header row {header_row}: {e}")
        return None


# ---------------------------------------------------------------------------
# Date shortcut helpers
# ---------------------------------------------------------------------------
def get_date_from_shortcut(shortcut):
    """Get a date or date range from a shortcut string."""
    today = date.today()
    
    if shortcut == 'today':
        return today, today
    elif shortcut == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif shortcut == 'this_week':
        start = today - timedelta(days=today.weekday())
        return start, today
    elif shortcut == 'this_month':
        start = date(today.year, today.month, 1)
        return start, today
    elif shortcut == 'last_month':
        first_day = date(today.year, today.month - 1 if today.month > 1 else 12, 1)
        if today.month == 1:
            last_day = date(today.year - 1, 12, 31)
        else:
            last_day = date(today.year, today.month, 1) - timedelta(days=1)
        return first_day, last_day
    elif shortcut == 'last_7_days':
        start = today - timedelta(days=7)
        return start, today
    elif shortcut == 'last_30_days':
        start = today - timedelta(days=30)
        return start, today
    elif shortcut == 'last_90_days':
        start = today - timedelta(days=90)
        return start, today
    elif shortcut == 'year_to_date':
        start = date(today.year, 1, 1)
        return start, today
    return None, None


def get_date_filter_description(shortcut):
    """Get a human-readable description of a date shortcut."""
    descriptions = {
        'exact': 'Exact Date',
        'today': 'Today',
        'yesterday': 'Yesterday',
        'this_week': 'This Week (Mon-Today)',
        'this_month': 'This Month (1st-Today)',
        'last_month': 'Last Month',
        'last_7_days': 'Last 7 Days',
        'last_30_days': 'Last 30 Days',
        'last_90_days': 'Last 90 Days',
        'year_to_date': 'Year to Date'
    }
    return descriptions.get(shortcut, shortcut)


def apply_date_shortcut_to_filter(df, column, operation, shortcut, value=None):
    """Apply a date shortcut to filter a dataframe column."""
    if shortcut == 'exact':
        if value is None:
            return df
        return MultiSheetProcessor.filter_rows(df, column, operation, value)
    
    start_date, end_date = get_date_from_shortcut(shortcut)
    if start_date is None or end_date is None:
        return df
    
    # Convert to datetime for comparison
    start_datetime = pd.Timestamp(start_date)
    end_datetime = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    
    # Get the column as datetime
    col_series = pd.to_datetime(df[column], errors='coerce')
    
    # Apply filter based on operation
    if operation in ['==', 'contains']:
        mask = (col_series >= start_datetime) & (col_series <= end_datetime)
    elif operation == '!=':
        mask = ~((col_series >= start_datetime) & (col_series <= end_datetime))
    elif operation == '>':
        mask = col_series > end_datetime
    elif operation == '<':
        mask = col_series < start_datetime
    elif operation == '>=':
        mask = col_series >= start_datetime
    elif operation == '<=':
        mask = col_series <= end_datetime
    else:
        return df
    
    return df[mask]

# ---------------------------------------------------------------------------
# Password-protected file handling
# ---------------------------------------------------------------------------
def is_excel_password_protected(file_bytes):
    """Check if an Excel file is password protected by attempting to read it."""
    try:
        bio = io.BytesIO(file_bytes)
        try:
            wb = openpyxl.load_workbook(bio, read_only=True)
            return False
        except Exception as e:
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                return True
            return False
    except Exception:
        return False


def decrypt_excel_file(file_bytes, password):
    """Decrypt a password-protected Excel file using msoffcrypto."""
    if not HAS_MSOFFCRYPTO:
        st.error("msoffcrypto is not installed. Please install it: pip install msoffcrypto-tool")
        return None
    
    try:
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_input.write(file_bytes)
        temp_input.close()
        
        temp_output = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_output.close()
        
        with open(temp_input.name, 'rb') as f:
            file = msoffcrypto.OfficeFile(f)
            file.load_key(password=password)
            
            with open(temp_output.name, 'wb') as f_out:
                file.decrypt(f_out)
        
        with open(temp_output.name, 'rb') as f:
            decrypted_bytes = f.read()
        
        try:
            os.unlink(temp_input.name)
            os.unlink(temp_output.name)
        except:
            pass
        
        return decrypted_bytes
    except Exception as e:
        st.error(f"Failed to decrypt file: {str(e)}")
        return None


def decrypt_excel_file_with_password(file_bytes, password):
    """Try to decrypt an Excel file with a given password."""
    try:
        if not is_excel_password_protected(file_bytes):
            return file_bytes
        
        decrypted = decrypt_excel_file(file_bytes, password)
        if decrypted is not None:
            return decrypted
        return None
    except Exception as e:
        return None


def load_excel_with_password(file_bytes, password=None):
    """Load an Excel file, handling password protection if present."""
    try:
        if is_excel_password_protected(file_bytes):
            if password is None:
                return None, "password_required"
            
            decrypted_bytes = decrypt_excel_file(file_bytes, password)
            if decrypted_bytes is None:
                return None, "wrong_password"
            file_bytes = decrypted_bytes
        
        # Get all sheet names
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        sheets = {}
        
        for sheet_name in xl.sheet_names:
            # Load without header first to detect header row
            raw_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None, nrows=20, dtype=str)
            detected_header_row = detect_header_row(raw_df)
            
            # Load with detected header row
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=detected_header_row)
            df.columns = dedupe_column_names(df.columns)
            for col in df.columns:
                converted = try_convert_datetime_column(df[col])
                if converted is not None:
                    df[col] = converted
            sheets[sheet_name] = df
        
        return sheets, "success"
    except Exception as e:
        return None, f"error: {str(e)}"


def load_csv_with_password(file_bytes, password=None):
    """Load a CSV file (CSV files don't typically have passwords)."""
    try:
        # Detect header row in CSV
        raw_df = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=20, dtype=str)
        detected_header_row = detect_header_row(raw_df)
        
        df = pd.read_csv(io.BytesIO(file_bytes), header=detected_header_row)
        df.columns = dedupe_column_names(df.columns)
        for col in df.columns:
            converted = try_convert_datetime_column(df[col])
            if converted is not None:
                df[col] = converted
        return {"Sheet1": df}, "success"
    except Exception as e:
        return None, f"error: {str(e)}"

# ---------------------------------------------------------------------------
# Formula engine helpers (full set from oldwiz)
# ---------------------------------------------------------------------------
def _elementwise(scalar_func):
    def wrapper(x, *args, **kwargs):
        if isinstance(x, pd.Series):
            return x.apply(lambda v: scalar_func(v, *args, **kwargs))
        return scalar_func(x, *args, **kwargs)
    return wrapper


def _s(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x)


def _f_upper(x):
    return x.str.upper() if hasattr(x, "str") else _s(x).upper()


def _f_lower(x):
    return x.str.lower() if hasattr(x, "str") else _s(x).lower()


def _f_len(x):
    if hasattr(x, "str"):
        return x.str.len()
    return pd.Series([len(str(v)) for v in x]) if hasattr(x, "__iter__") else len(str(x))


def _f_proper(x):
    return _elementwise(lambda v: _s(v).title())(x)


def _f_trim(x):
    return _elementwise(lambda v: re.sub(r"\s+", " ", _s(v)).strip())(x)


def _f_clean(x):
    return _elementwise(lambda v: "".join(c for c in _s(v) if c.isprintable()))(x)


def _f_left(x, n):
    return _elementwise(lambda v: _s(v)[: int(n)])(x)


def _f_right(x, n):
    n = int(n)
    return _elementwise(lambda v: _s(v)[-n:] if n > 0 else "")(x)


def _f_mid(x, start, n):
    return _elementwise(lambda v: _s(v)[int(start) - 1: int(start) - 1 + int(n)])(x)


def _f_concat(*args):
    series_args = [a for a in args if isinstance(a, pd.Series)]
    if not series_args:
        return "".join(_s(a) for a in args)
    length = len(series_args[0])
    idx = series_args[0].index
    out = []
    for i in range(length):
        parts = []
        for a in args:
            parts.append(_s(a.iloc[i]) if isinstance(a, pd.Series) else _s(a))
        out.append("".join(parts))
    return pd.Series(out, index=idx)


def _f_textjoin(delimiter, ignore_empty, *args):
    series_args = [a for a in args if isinstance(a, pd.Series)]
    if not series_args:
        parts = [_s(a) for a in args if not (ignore_empty and _s(a) == "")]
        return _s(delimiter).join(parts)
    length = len(series_args[0])
    idx = series_args[0].index
    out = []
    for i in range(length):
        parts = []
        for a in args:
            val = _s(a.iloc[i]) if isinstance(a, pd.Series) else _s(a)
            if ignore_empty and val == "":
                continue
            parts.append(val)
        out.append(_s(delimiter).join(parts))
    return pd.Series(out, index=idx)


def _f_substitute(x, old, new, occurrence=None):
    def _one(v):
        v = _s(v)
        if occurrence is None:
            return v.replace(_s(old), _s(new))
        parts = v.split(_s(old))
        if len(parts) <= int(occurrence):
            return v
        rebuilt = _s(old).join(parts[: int(occurrence)]) + _s(new) + _s(old).join(parts[int(occurrence):])
        return rebuilt
    return _elementwise(_one)(x)


def _f_replace(x, start, n, new_text):
    def _one(v):
        v = _s(v)
        start_i = int(start) - 1
        return v[:start_i] + _s(new_text) + v[start_i + int(n):]
    return _elementwise(_one)(x)


def _f_find(find_text, within_text, start=1):
    def _one(v):
        pos = _s(v).find(_s(find_text), int(start) - 1)
        return pos + 1 if pos >= 0 else np.nan
    return _elementwise(_one)(within_text)


def _f_search(find_text, within_text, start=1):
    def _one(v):
        pos = _s(v).lower().find(_s(find_text).lower(), int(start) - 1)
        return pos + 1 if pos >= 0 else np.nan
    return _elementwise(_one)(within_text)


def _f_rept(x, n):
    return _elementwise(lambda v: _s(v) * int(n))(x)


def _f_exact(a, b):
    if isinstance(a, pd.Series) or isinstance(b, pd.Series):
        base = a if isinstance(a, pd.Series) else b
        other = b if isinstance(a, pd.Series) else a
        if isinstance(a, pd.Series) and isinstance(b, pd.Series):
            return (a.astype(str) == b.astype(str))
        return base.astype(str) == _s(other)
    return _s(a) == _s(b)


def _f_value(x):
    def _one(v):
        try:
            return float(str(v).replace(",", "").replace("$", "").replace("%", ""))
        except Exception:
            return np.nan
    return _elementwise(_one)(x)


def _f_text(x, fmt=""):
    def _one(v):
        try:
            if isinstance(v, (pd.Timestamp,)) or "y" in fmt.lower() or "m" in fmt.lower() and "d" in fmt.lower():
                ts = pd.to_datetime(v, errors="coerce")
                if pd.notna(ts):
                    py_fmt = (fmt.replace("YYYY", "%Y").replace("yyyy", "%Y")
                              .replace("MM", "%m").replace("DD", "%d")
                              .replace("HH", "%H").replace("mm", "%M").replace("SS", "%S"))
                    return ts.strftime(py_fmt) if "%" in py_fmt else str(v)
            if fmt in ("0", "0.00", "#,##0", "#,##0.00"):
                decimals = fmt.count("0") - fmt.split(".")[0].count("0") if "." in fmt else 0
                num = float(v)
                return f"{num:,.{decimals}f}" if "," in fmt else f"{num:.{decimals}f}"
            return str(v)
        except Exception:
            return str(v)
    return _elementwise(_one)(x)


def _agg_over_args(args, agg):
    values = []
    for a in args:
        if isinstance(a, pd.Series):
            values.extend(pd.to_numeric(a, errors="coerce").dropna().tolist())
        elif a is not None and not (isinstance(a, float) and pd.isna(a)):
            try:
                values.append(float(a))
            except Exception:
                pass
    if not values:
        return 0 if agg in ("sum", "count") else np.nan
    arr = np.array(values, dtype=float)
    return {
        "sum": np.sum, "average": np.mean, "median": np.median,
        "max": np.max, "min": np.min, "count": len, "stdev": (lambda a: np.std(a, ddof=1) if len(a) > 1 else 0.0),
        "var": (lambda a: np.var(a, ddof=1) if len(a) > 1 else 0.0),
    }[agg](arr)


def _f_sum(*args):
    return _agg_over_args(args, "sum")


def _f_average(*args):
    return _agg_over_args(args, "average")


def _f_median(*args):
    return _agg_over_args(args, "median")


def _f_count(*args):
    values = []
    for a in args:
        if isinstance(a, pd.Series):
            values.extend(pd.to_numeric(a, errors="coerce").dropna().tolist())
        elif a is not None and not (isinstance(a, float) and pd.isna(a)):
            try:
                values.append(float(a))
            except Exception:
                pass
    return len(values)


def _f_counta(*args):
    n = 0
    for a in args:
        if isinstance(a, pd.Series):
            n += a.notna().sum()
        elif a is not None and not (isinstance(a, float) and pd.isna(a)):
            n += 1
    return n


def _f_max(*args):
    return _agg_over_args(args, "max")


def _f_min(*args):
    return _agg_over_args(args, "min")


def _f_stdev(*args):
    return _agg_over_args(args, "stdev")


def _f_var(*args):
    return _agg_over_args(args, "var")


def _f_countif(rng, criteria):
    mask = _f_criteria_mask(rng, criteria)
    return int(mask.sum())


def _f_sumif(rng, criteria, sum_range=None):
    mask = _f_criteria_mask(rng, criteria)
    target = sum_range if sum_range is not None else rng
    return float(pd.to_numeric(target[mask], errors="coerce").fillna(0).sum())


def _f_averageif(rng, criteria, avg_range=None):
    mask = _f_criteria_mask(rng, criteria)
    target = avg_range if avg_range is not None else rng
    vals = pd.to_numeric(target[mask], errors="coerce").dropna()
    return float(vals.mean()) if len(vals) else np.nan


def _f_criteria_mask(rng, criteria):
    if not isinstance(rng, pd.Series):
        raise Exception("COUNTIF/SUMIF/AVERAGEIF need a column (e.g. [Column]) as the range")
    crit = str(criteria).strip()
    m = re.match(r'^(>=|<=|<>|>|<|=)\s*(.+)$', crit)
    if m:
        op, val = m.group(1), m.group(2).strip()
        try:
            num_val = float(val)
            series_num = pd.to_numeric(rng, errors="coerce")
            if op == ">": return series_num > num_val
            if op == "<": return series_num < num_val
            if op == ">=": return series_num >= num_val
            if op == "<=": return series_num <= num_val
            if op == "<>": return series_num != num_val
            if op == "=": return series_num == num_val
        except ValueError:
            text_val = val
            if op == "<>": return rng.astype(str) != text_val
            if op == "=": return rng.astype(str) == text_val
            raise Exception(f"Non-numeric comparison '{crit}' isn't supported for text columns")
    if "*" in crit or "?" in crit:
        pattern = "^" + re.escape(crit).replace(r"\*", ".*").replace(r"\?", ".") + "$"
        return rng.astype(str).str.match(pattern, case=False, na=False)
    return rng.astype(str).str.lower() == crit.lower()


def _f_rank(value, rng, ascending=False):
    if not isinstance(rng, pd.Series):
        raise Exception("RANK needs a column (e.g. [Column]) as the range")
    ranked = rng.rank(ascending=bool(ascending), method="min")
    if isinstance(value, pd.Series):
        return value.map(lambda v: ranked[rng == v].iloc[0] if (rng == v).any() else np.nan)
    matches = ranked[rng == value]
    return matches.iloc[0] if len(matches) else np.nan


def _f_round(x, n=0):
    return _elementwise(lambda v: round(float(v), int(n)))(x)


def _f_roundup(x, n=0):
    factor = 10 ** int(n)
    return _elementwise(lambda v: np.ceil(float(v) * factor) / factor)(x)


def _f_rounddown(x, n=0):
    factor = 10 ** int(n)
    return _elementwise(lambda v: np.floor(float(v) * factor) / factor)(x)


def _f_mod(x, divisor):
    return _elementwise(lambda v: float(v) % float(divisor))(x)


def _f_power(x, n):
    return _elementwise(lambda v: float(v) ** float(n))(x)


def _f_sqrt(x):
    return _elementwise(lambda v: np.sqrt(float(v)))(x)


def _f_int(x):
    return _elementwise(lambda v: int(np.floor(float(v))))(x)


def _f_ceiling(x, significance=1):
    return _elementwise(lambda v: np.ceil(float(v) / float(significance)) * float(significance))(x)


def _f_floor(x, significance=1):
    return _elementwise(lambda v: np.floor(float(v) / float(significance)) * float(significance))(x)


def _f_and(*args):
    if any(isinstance(a, pd.Series) for a in args):
        result = None
        for a in args:
            cond = a if isinstance(a, pd.Series) else pd.Series([bool(a)])
            result = cond if result is None else (result & cond)
        return result
    return all(bool(a) for a in args)


def _f_or(*args):
    if any(isinstance(a, pd.Series) for a in args):
        result = None
        for a in args:
            cond = a if isinstance(a, pd.Series) else pd.Series([bool(a)])
            result = cond if result is None else (result | cond)
        return result
    return any(bool(a) for a in args)


def _f_not(x):
    return _elementwise(lambda v: not bool(v))(x)


def _f_if(cond, true_val, false_val):
    if isinstance(cond, pd.Series):
        return pd.Series(np.where(cond, true_val, false_val),
                          index=cond.index if hasattr(cond, "index") else None)
    return true_val if cond else false_val


def _f_iferror(x, fallback):
    def _bad(v):
        return v is None or (isinstance(v, float) and (pd.isna(v) or np.isinf(v)))
    try:
        if isinstance(x, pd.Series):
            return x.apply(lambda v: fallback if _bad(v) else v)
        return fallback if _bad(x) else x
    except Exception:
        return fallback


def _f_ifna(x, fallback):
    return _f_iferror(x, fallback)


def _f_isblank(x):
    return _elementwise(lambda v: v is None or (isinstance(v, float) and pd.isna(v)) or _s(v) == "")(x)


def _f_isnumber(x):
    def _one(v):
        try:
            float(v)
            return True
        except Exception:
            return False
    return _elementwise(_one)(x)


def _f_istext(x):
    return _elementwise(lambda v: isinstance(v, str))(x)


def _f_isna(x):
    return _elementwise(lambda v: v is None or (isinstance(v, float) and pd.isna(v)))(x)


def _f_today():
    return pd.Timestamp.now().normalize()


def _f_now():
    return pd.Timestamp.now()


def _f_date(y, m, d):
    def _one(y_, m_, d_):
        try:
            return pd.Timestamp(year=int(y_), month=int(m_), day=int(d_))
        except Exception:
            return pd.NaT
    if isinstance(y, pd.Series) or isinstance(m, pd.Series) or isinstance(d, pd.Series):
        length = len(y) if isinstance(y, pd.Series) else (len(m) if isinstance(m, pd.Series) else len(d))
        idx = (y.index if isinstance(y, pd.Series) else (m.index if isinstance(m, pd.Series) else d.index))
        out = []
        for i in range(length):
            yy = y.iloc[i] if isinstance(y, pd.Series) else y
            mm = m.iloc[i] if isinstance(m, pd.Series) else m
            dd = d.iloc[i] if isinstance(d, pd.Series) else d
            out.append(_one(yy, mm, dd))
        return pd.Series(out, index=idx)
    return _one(y, m, d)


def _f_year(x):
    return _elementwise(lambda v: pd.to_datetime(v, errors="coerce").year)(x)


def _f_month(x):
    return _elementwise(lambda v: pd.to_datetime(v, errors="coerce").month)(x)


def _f_day(x):
    return _elementwise(lambda v: pd.to_datetime(v, errors="coerce").day)(x)


def _f_weekday(x):
    return _elementwise(lambda v: pd.to_datetime(v, errors="coerce").weekday() + 1)(x)


def _f_eomonth(x, months=0):
    def _one(v):
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return pd.NaT
        return (ts + pd.DateOffset(months=int(months) + 1)).replace(day=1) - pd.Timedelta(days=1)
    return _elementwise(_one)(x)


def _f_datedif(start, end, unit="D"):
    def _one(s, e):
        s_ts, e_ts = pd.to_datetime(s, errors="coerce"), pd.to_datetime(e, errors="coerce")
        if pd.isna(s_ts) or pd.isna(e_ts):
            return np.nan
        unit_u = str(unit).upper()
        if unit_u == "D":
            return (e_ts - s_ts).days
        if unit_u == "M":
            return (e_ts.year - s_ts.year) * 12 + (e_ts.month - s_ts.month)
        if unit_u == "Y":
            return e_ts.year - s_ts.year
        return (e_ts - s_ts).days
    if isinstance(start, pd.Series) or isinstance(end, pd.Series):
        length = len(start) if isinstance(start, pd.Series) else len(end)
        idx = start.index if isinstance(start, pd.Series) else end.index
        out = []
        for i in range(length):
            s_ = start.iloc[i] if isinstance(start, pd.Series) else start
            e_ = end.iloc[i] if isinstance(end, pd.Series) else end
            out.append(_one(s_, e_))
        return pd.Series(out, index=idx)
    return _one(start, end)


class _SafeNP:
    nan = np.nan
    where = staticmethod(np.where)
    isnan = staticmethod(np.isnan)


class _SafePD:
    isna = staticmethod(pd.isna)
    notna = staticmethod(pd.notna)
    Timestamp = pd.Timestamp


SAFE_NAMESPACE_FUNCS = {
    "abs": abs,
    "round": _f_round,
    "min": min,
    "max": max,
    "upper": _f_upper,
    "lower": _f_lower,
    "len": _f_len,
    "np": _SafeNP,
    "pd": _SafePD,
    "UPPER": _f_upper, "LOWER": _f_lower, "PROPER": _f_proper, "TRIM": _f_trim,
    "CLEAN": _f_clean, "LEN": _f_len, "LEFT": _f_left, "RIGHT": _f_right, "MID": _f_mid,
    "CONCAT": _f_concat, "CONCATENATE": _f_concat, "TEXTJOIN": _f_textjoin,
    "SUBSTITUTE": _f_substitute, "REPLACE": _f_replace, "FIND": _f_find, "SEARCH": _f_search,
    "REPT": _f_rept, "EXACT": _f_exact, "VALUE": _f_value, "TEXT": _f_text,
    "SUM": _f_sum, "AVERAGE": _f_average, "MEDIAN": _f_median, "COUNT": _f_count,
    "COUNTA": _f_counta, "STDEV": _f_stdev, "VAR": _f_var, "MAX": _f_max, "MIN": _f_min,
    "ROUND": _f_round, "ROUNDUP": _f_roundup, "ROUNDDOWN": _f_rounddown, "ABS": abs,
    "MOD": _f_mod, "POWER": _f_power, "SQRT": _f_sqrt, "INT": _f_int,
    "CEILING": _f_ceiling, "FLOOR": _f_floor, "RANK": _f_rank,
    "COUNTIF": _f_countif, "SUMIF": _f_sumif, "AVERAGEIF": _f_averageif,
    "AND": _f_and, "OR": _f_or, "NOT": _f_not, "IF": _f_if,
    "IFERROR": _f_iferror, "IFNA": _f_ifna,
    "ISBLANK": _f_isblank, "ISNUMBER": _f_isnumber, "ISTEXT": _f_istext, "ISNA": _f_isna,
    "TODAY": _f_today, "NOW": _f_now, "DATE": _f_date, "YEAR": _f_year, "MONTH": _f_month,
    "DAY": _f_day, "WEEKDAY": _f_weekday, "EOMONTH": _f_eomonth, "DATEDIF": _f_datedif,
}

_FORMULA_BLOCKLIST_PATTERN = re.compile(
    r"__|\bimport\b|\bexec\b|\beval\b|\bopen\b|\bcompile\b|\bgetattr\b|"
    r"\bsetattr\b|\bdelattr\b|\bglobals\b|\blocals\b|\bvars\b|\bos\.|\bsys\.|\bsubprocess\b",
    re.IGNORECASE,
)


def _assert_formula_is_safe(formula):
    if _FORMULA_BLOCKLIST_PATTERN.search(formula):
        raise Exception(
            "This formula contains characters/keywords that aren't allowed "
            "(e.g. '__', 'import', 'os.', 'exec'). Please use plain column "
            "references, arithmetic, and the supported functions only."
        )


def _replace_columns(expr, df):
    def repl(match):
        col_name = match.group(1)
        escaped = col_name.replace("\\", "\\\\").replace("'", "\\'")
        return f"df['{escaped}']"
    return re.sub(r"\[([^\]]+)\]", repl, expr)


def _xlookup(lookup_value, ref_source_name, lookup_column, return_column, if_not_found=None, sources=None):
    sources = sources or {}
    ref_df = sources.get(ref_source_name)
    if ref_df is None:
        raise Exception(f"Reference source '{ref_source_name}' is not loaded")
    if lookup_column not in ref_df.columns:
        raise Exception(f"'{lookup_column}' not found in reference source '{ref_source_name}'")
    if return_column not in ref_df.columns:
        raise Exception(f"'{return_column}' not found in reference source '{ref_source_name}'")

    mapping = dict(zip(ref_df[lookup_column], ref_df[return_column]))

    if hasattr(lookup_value, "map"):
        result = lookup_value.map(mapping)
        if if_not_found is not None:
            result = result.fillna(if_not_found)
        return result
    return mapping.get(lookup_value, if_not_found)


def detect_referenced_source_names(text, known_source_names):
    found = []
    for name in known_source_names:
        if f"'{name}'" in text or f'"{name}"' in text:
            found.append(name)
    return found


def try_convert_datetime_column(series):
    if pd.api.types.is_numeric_dtype(series):
        return None
    if pd.api.types.is_datetime64_any_dtype(series):
        return series

    non_null = series.dropna()
    if len(non_null) == 0:
        return None

    numeric_like = pd.to_numeric(non_null.astype(str).str.replace(',', '', regex=False), errors='coerce')
    if numeric_like.notna().mean() >= 0.9:
        return None

    try:
        parsed = pd.to_datetime(non_null, errors='coerce')
    except Exception:
        return None

    parse_rate = parsed.notna().mean()
    if parse_rate >= 0.9:
        return pd.to_datetime(series, errors='coerce')
    return None


def dedupe_column_names(columns):
    seen = {}
    result = []
    for i, col in enumerate(columns):
        name = str(col).strip()
        if name == "" or name.lower() == "nan":
            name = f"Column_{i+1}"
        if name in seen:
            seen[name] += 1
            new_name = f"{name}_{seen[name]}"
            while new_name in seen:
                seen[name] += 1
                new_name = f"{name}_{seen[name]}"
            seen[new_name] = 0
            result.append(new_name)
        else:
            seen[name] = 0
            result.append(name)
    return result


def get_unique_values_sorted(df, column, max_values=1000):
    try:
        values = list(df[column].dropna().unique())
        try:
            values = sorted(values)
        except TypeError:
            values = sorted(values, key=lambda v: str(v))
        if len(values) > max_values:
            return values[:max_values], True
        return values, False
    except Exception:
        return [], False


def read_raw_preview_bytes(file_bytes, ext, sheet_name=None, nrows=15):
    bio = io.BytesIO(file_bytes)
    if ext == ".csv":
        return pd.read_csv(bio, header=None, nrows=nrows, dtype=str)
    return pd.read_excel(bio, sheet_name=sheet_name, header=None, nrows=nrows, dtype=str)


def analyze_header_ambiguity(raw_preview):
    """Analyze if the header row is ambiguous (not in row 0)."""
    if raw_preview is None or len(raw_preview) == 0:
        return False, 0
    
    # Check if row 0 looks like a header
    row0 = raw_preview.iloc[0]
    row0_has_text = False
    text_count = 0
    for v in row0.dropna():
        try:
            float(str(v).replace(',', '').replace('$', '').replace('%', ''))
        except (ValueError, TypeError):
            text_count += 1
    
    if len(row0.dropna()) > 0:
        row0_has_text = text_count / len(row0.dropna()) > 0.4
    
    # If row 0 already looks like a header, it's not ambiguous
    if row0_has_text and len(row0.dropna()) >= 2:
        return False, 0
    
    # Check other rows for better candidates
    best_row = detect_header_row(raw_preview)
    if best_row > 0:
        return True, best_row
    
    return False, 0


def columns_look_unnamed_heavy(columns, threshold=0.3):
    cols = [str(c) for c in columns]
    if not cols:
        return False
    unnamed = sum(1 for c in cols if re.match(r"^Unnamed:\s*\d+$", c) or re.match(r"^Column_\d+$", c))
    return (unnamed / len(cols)) >= threshold


def apply_output_format(series, format_cfg):
    if not format_cfg or format_cfg.get('type') in (None, 'none'):
        return series

    ftype = format_cfg.get('type')

    try:
        if ftype == 'date':
            pattern = format_cfg.get('pattern', '%Y-%m-%d')
            parsed = pd.to_datetime(series, errors='coerce')
            formatted = parsed.dt.strftime(pattern)
            return formatted.where(parsed.notna(), series)

        elif ftype == 'number':
            decimals = int(format_cfg.get('decimals', 2))
            use_thousands = bool(format_cfg.get('thousands', False))
            prefix = format_cfg.get('prefix', '') or ''
            suffix = format_cfg.get('suffix', '') or ''

            def _fmt_num(v):
                if pd.isna(v):
                    return v
                try:
                    x = float(str(v).replace(',', ''))
                except (TypeError, ValueError):
                    return v
                spec = f",.{decimals}f" if use_thousands else f".{decimals}f"
                return f"{prefix}{format(x, spec)}{suffix}"

            return series.map(_fmt_num)

        elif ftype == 'text':
            case = format_cfg.get('case', 'as_is')
            trim = format_cfg.get('trim', False)

            def _fmt_text(v):
                if pd.isna(v):
                    return v
                s = str(v)
                if trim:
                    s = s.strip()
                if case == 'upper':
                    s = s.upper()
                elif case == 'lower':
                    s = s.lower()
                elif case == 'title':
                    s = s.title()
                return s

            return series.map(_fmt_text)

    except Exception:
        return series

    return series

# ---------------------------------------------------------------------------
# MultiSheetProcessor Class
# ---------------------------------------------------------------------------
class MultiSheetProcessor:
    """Handles multiple sheets and files with references"""

    @staticmethod
    def load_multi_sheet_file_with_header_detection(file_bytes, ext):
        """Load all sheets from an Excel file with header detection."""
        try:
            xl = pd.ExcelFile(io.BytesIO(file_bytes))
            sheets = {}
            pending_review = []
            
            for sheet_name in xl.sheet_names:
                # Read raw preview for header detection
                raw_preview = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, 
                                            header=None, nrows=20, dtype=str)
                
                # Check if header needs review
                ambiguous, suggested_row = analyze_header_ambiguity(raw_preview)
                
                if ambiguous:
                    pending_review.append({
                        'file_name': 'current_file',
                        'sheet_name': sheet_name,
                        'raw_preview': raw_preview,
                        'suggested_row': suggested_row
                    })
                else:
                    # Load with detected or default header row
                    detected_row = detect_header_row(raw_preview)
                    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=detected_row)
                    df.columns = dedupe_column_names(df.columns)
                    for col in df.columns:
                        converted = try_convert_datetime_column(df[col])
                        if converted is not None:
                            df[col] = converted
                    sheets[sheet_name] = df
            
            return sheets, pending_review
        except Exception as e:
            st.error(f"Error loading multi-sheet file: {e}")
            return None, []

    @staticmethod
    def load_single_sheet_with_header_detection(file_bytes, ext, sheet_name=None):
        """Load a single sheet with header detection."""
        try:
            # Read raw preview
            if ext == ".csv":
                raw_preview = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=20, dtype=str)
            else:
                raw_preview = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, 
                                            header=None, nrows=20, dtype=str)
            
            # Detect header row
            detected_row = detect_header_row(raw_preview)
            
            # Load with detected header
            if ext == ".csv":
                df = pd.read_csv(io.BytesIO(file_bytes), header=detected_row)
            else:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=detected_row)
            
            df.columns = dedupe_column_names(df.columns)
            for col in df.columns:
                converted = try_convert_datetime_column(df[col])
                if converted is not None:
                    df[col] = converted
            return df
        except Exception as e:
            st.error(f"Error loading sheet: {e}")
            return None

    @staticmethod
    def load_multi_sheet_file(uploaded_file):
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            sheets = {}

            for sheet_name in excel_file.sheet_names:
                # Try to detect header row
                raw_preview = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, nrows=20, dtype=str)
                detected_row = detect_header_row(raw_preview)
                
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=detected_row)
                df.columns = dedupe_column_names(df.columns)

                for col in df.columns:
                    converted = try_convert_datetime_column(df[col])
                    if converted is not None:
                        df[col] = converted

                sheets[sheet_name] = df

            return sheets

        except Exception as e:
            st.error(f"Error loading Excel file: {e}")
            return None

    @staticmethod
    def load_excel_sheet_with_header(file_bytes, sheet_name, header_row=0):
        bio = io.BytesIO(file_bytes)
        df = pd.read_excel(bio, sheet_name=sheet_name, header=header_row)
        df.columns = dedupe_column_names(df.columns)
        for col in df.columns:
            converted = try_convert_datetime_column(df[col])
            if converted is not None:
                df[col] = converted
        return df

    @staticmethod
    def load_csv_with_header(file_bytes, header_row=0):
        bio = io.BytesIO(file_bytes)
        df = pd.read_csv(bio, header=header_row)
        df.columns = dedupe_column_names(df.columns)
        for col in df.columns:
            converted = try_convert_datetime_column(df[col])
            if converted is not None:
                df[col] = converted
        return df

    @staticmethod
    def load_single_sheet(uploaded_file):
        """Load a single sheet from a file (for backward compatibility)."""
        try:
            ext = Path(uploaded_file.name).suffix.lower()

            if ext == ".csv":
                df = pd.read_csv(uploaded_file)
                df.columns = dedupe_column_names(df.columns)

                for col in df.columns:
                    converted = try_convert_datetime_column(df[col])
                    if converted is not None:
                        df[col] = converted

                return {"Sheet1": df}

            elif ext in [".xlsx", ".xls"]:
                return MultiSheetProcessor.load_multi_sheet_file(uploaded_file)

            st.error("Unsupported file type")
            return None

        except Exception as e:
            st.error(f"Error loading file: {e}")
            return None

    @staticmethod
    def get_filter_mask(df, column, operation, value):
        if column not in df.columns:
            st.error(f"Column '{column}' not found!")
            return None

        series = df[column]

        converted_dates = try_convert_datetime_column(series)
        is_date = converted_dates is not None
        if is_date:
            series = converted_dates

        if is_date:
            try:
                compare = pd.to_datetime(value, errors='coerce')
                if pd.isna(compare):
                    st.error("Invalid date format. Use YYYY-MM-DD or similar.")
                    return None
            except Exception:
                st.error("Invalid date format. Use YYYY-MM-DD or similar.")
                return None

            series_normalized = series.dt.normalize()
            compare_normalized = compare.normalize()

            if operation == "==":
                return series_normalized == compare_normalized
            elif operation == "!=":
                return series_normalized != compare_normalized
            elif operation == ">":
                return series > compare
            elif operation == "<":
                return series < compare
            elif operation == ">=":
                return series >= compare
            elif operation == "<=":
                return series <= compare
            elif operation == "contains":
                return series.dt.strftime("%Y-%m-%d").str.contains(str(value), case=False, na=False)
            return None

        if pd.api.types.is_numeric_dtype(series):
            try:
                num_value = float(value)
            except ValueError:
                st.error("Please enter a numeric value for this column")
                return None

            if operation == "==":
                return series == num_value
            elif operation == "!=":
                return series != num_value
            elif operation == ">":
                return series > num_value
            elif operation == "<":
                return series < num_value
            elif operation == ">=":
                return series >= num_value
            elif operation == "<=":
                return series <= num_value
            return None

        series = series.astype(str).str.strip()
        value_str = str(value).strip()

        if operation == "contains":
            return series.str.contains(value_str, case=False, na=False)
        elif operation == "==":
            return series.str.lower() == value_str.lower()
        elif operation == "!=":
            return series.str.lower() != value_str.lower()
        elif operation == ">":
            return series > value_str
        elif operation == "<":
            return series < value_str
        elif operation == ">=":
            return series >= value_str
        elif operation == "<=":
            return series <= value_str
        return None

    @staticmethod
    def filter_rows(df, column, operation, value):
        try:
            mask = MultiSheetProcessor.get_filter_mask(df, column, operation, value)
            if mask is None:
                return df
            return df[mask]
        except Exception as e:
            st.error(f"Error filtering: {e}")
            return df

    @staticmethod
    def filter_multiple_rows(df, conditions, logic='AND'):
        try:
            masks = []
            for condition in conditions:
                column = condition.get('column')
                operation = condition.get('operation')
                value = condition.get('value')
                if column and operation and value is not None:
                    mask = MultiSheetProcessor.get_filter_mask(df, column, operation, value)
                    if mask is not None:
                        masks.append(mask)

            if not masks:
                return df

            combined = masks[0]
            for m in masks[1:]:
                combined = (combined | m) if logic == 'OR' else (combined & m)

            return df[combined]
        except Exception as e:
            st.error(f"Error applying multiple filters: {e}")
            return df

    @staticmethod
    def add_calculated_column(df, new_column, formula, all_sources=None):
        try:
            result = MultiSheetProcessor.evaluate_formula(df, formula, all_sources)
            df = df.copy()
            df[new_column] = result
            return df
        except Exception as e:
            st.error(f"Error adding column: {e}")
            return df

    @staticmethod
    def evaluate_formula(df, formula, all_sources=None):
        try:
            _assert_formula_is_safe(formula)

            formula = formula.replace("<>", "!=")
            formula = re.sub(r'(?<![<>=!])=(?!=)', '==', formula)

            namespace = dict(SAFE_NAMESPACE_FUNCS)
            namespace["df"] = df
            namespace["XLOOKUP"] = lambda lookup_value, ref_name, lookup_col, return_col, if_not_found=None: \
                _xlookup(lookup_value, ref_name, lookup_col, return_col, if_not_found, sources=all_sources)
            namespace["VLOOKUP"] = namespace["XLOOKUP"]

            if_match = re.search(
                r'\bIF\b\s+(.+?)\s+\bTHEN\b\s+(.+?)\s+\bELSE\b\s+(.+?)\s+\bEND\b',
                formula, re.IGNORECASE | re.DOTALL
            )
            if if_match:
                condition = _replace_columns(if_match.group(1).strip(), df)
                true_val = _replace_columns(if_match.group(2).strip(), df)
                false_val = _replace_columns(if_match.group(3).strip(), df)

                cond_result = eval(condition, {"__builtins__": {}}, namespace)
                true_result = eval(true_val, {"__builtins__": {}}, namespace)
                false_result = eval(false_val, {"__builtins__": {}}, namespace)

                return pd.Series(
                    np.where(cond_result, true_result, false_result),
                    index=df.index
                )

            expr = _replace_columns(formula, df)
            result = eval(expr, {"__builtins__": {}}, namespace)
            return result
        except Exception as e:
            missing = [m for m in re.findall(r"\[([^\]]+)\]", formula) if m not in df.columns]
            if missing:
                raise Exception(
                    f"Error evaluating formula: column(s) not found in the current data: {missing}"
                )
            raise Exception(f"Error evaluating formula: {e}")

    @staticmethod
    def group_by(df, group_column, agg_column, agg_function):
        try:
            return df.groupby(group_column)[agg_column].agg(agg_function).reset_index()
        except Exception as e:
            st.error(f"Error grouping: {e}")
            return df

    @staticmethod
    def sort_data(df, column, ascending=True):
        try:
            return df.sort_values(by=column, ascending=ascending)
        except Exception as e:
            st.error(f"Error sorting: {e}")
            return df

    @staticmethod
    def rename_column(df, old_name, new_name):
        try:
            return df.rename(columns={old_name: new_name})
        except Exception as e:
            st.error(f"Error renaming: {e}")
            return df

    @staticmethod
    def remove_duplicates(df, subset=None, keep='first'):
        try:
            keep_arg = False if keep == 'drop_all' else keep
            return df.drop_duplicates(subset=subset, keep=keep_arg)
        except Exception as e:
            st.error(f"Error removing duplicates: {e}")
            return df

    @staticmethod
    def handle_missing_values(df, column, method, value=None):
        try:
            df = df.copy()
            cols = list(df.columns) if column in (None, '') else [column]
            if method == 'drop_rows':
                return df.dropna(subset=cols)
            for c in cols:
                if method == 'fill_static':
                    df[c] = df[c].fillna(value)
                elif method == 'fill_mean':
                    if pd.api.types.is_numeric_dtype(df[c]):
                        df[c] = df[c].fillna(df[c].mean())
                    else:
                        st.warning(f"Column '{c}' isn't numeric — skipped mean fill.")
                elif method == 'fill_median':
                    if pd.api.types.is_numeric_dtype(df[c]):
                        df[c] = df[c].fillna(df[c].median())
                    else:
                        st.warning(f"Column '{c}' isn't numeric — skipped median fill.")
                elif method == 'fill_mode':
                    mode_vals = df[c].mode(dropna=True)
                    if len(mode_vals) > 0:
                        df[c] = df[c].fillna(mode_vals.iloc[0])
                elif method == 'fill_forward':
                    df[c] = df[c].ffill()
                elif method == 'fill_backward':
                    df[c] = df[c].bfill()
            return df
        except Exception as e:
            st.error(f"Error handling missing values: {e}")
            return df

    @staticmethod
    def change_data_type(df, column, target_type):
        try:
            df = df.copy()
            original = df[column]

            if target_type == 'text':
                df[column] = original.astype(str).where(original.notna(), original)

            elif target_type in ('integer', 'float'):
                converted = pd.to_numeric(original, errors='coerce')
                failed = converted.isna() & original.notna()
                if failed.any():
                    st.warning(f"⚠️ {int(failed.sum())} value(s) in '{column}' couldn't convert to a "
                               f"number and became blank.")
                df[column] = converted.astype('Int64') if target_type == 'integer' else converted

            elif target_type == 'date':
                converted = pd.to_datetime(original, errors='coerce')
                failed = converted.isna() & original.notna()
                if failed.any():
                    st.warning(f"⚠️ {int(failed.sum())} value(s) in '{column}' couldn't convert to a "
                               f"date and became blank.")
                df[column] = converted

            elif target_type == 'boolean':
                truthy = {'true', 'yes', 'y', '1', '1.0'}
                falsy = {'false', 'no', 'n', '0', '0.0'}

                def _to_bool(v):
                    if pd.isna(v):
                        return np.nan
                    s = str(v).strip().lower()
                    if s in truthy:
                        return True
                    if s in falsy:
                        return False
                    return np.nan

                converted = original.map(_to_bool)
                failed = converted.isna() & original.notna()
                if failed.any():
                    st.warning(f"⚠️ {int(failed.sum())} value(s) in '{column}' didn't look like "
                               f"yes/no values and became blank.")
                df[column] = converted

            return df
        except Exception as e:
            st.error(f"Error changing data type: {e}")
            return df

    @staticmethod
    def find_replace(df, column, find_text, replace_text, match_case=False, whole_cell=False):
        try:
            df = df.copy()
            if column in (None, ''):
                cols = list(df.select_dtypes(include=['object']).columns)
            else:
                cols = [column]

            for c in cols:
                series = df[c].astype(str)
                if whole_cell:
                    if match_case:
                        mask = series == find_text
                    else:
                        mask = series.str.lower() == str(find_text).lower()
                    df[c] = df[c].where(~mask, replace_text)
                else:
                    df[c] = series.str.replace(find_text, replace_text, case=match_case, regex=False)
            return df
        except Exception as e:
            st.error(f"Error in find & replace: {e}")
            return df

    @staticmethod
    def split_column(df, column, delimiter, new_names=None):
        try:
            df = df.copy()
            split_df = df[column].astype(str).str.split(delimiter, expand=True)
            n = split_df.shape[1]
            if new_names and len(new_names) == n:
                split_df.columns = new_names
            else:
                split_df.columns = [f"{column}_{i+1}" for i in range(n)]

            for c in split_df.columns:
                if c in df.columns and c != column:
                    st.warning(f"⚠️ Column '{c}' already existed and was overwritten by the split result.")
                df[c] = split_df[c]
            return df
        except Exception as e:
            st.error(f"Error splitting column: {e}")
            return df

    @staticmethod
    def drop_column(df, column):
        try:
            return df.drop(columns=[column])
        except Exception as e:
            st.error(f"Error dropping column: {e}")
            return df

    @staticmethod
    def select_columns(df, columns):
        try:
            existing = [c for c in columns if c in df.columns]
            missing = [c for c in columns if c not in df.columns]
            if missing:
                st.warning(f"⚠️ These selected columns weren't found and were skipped: {missing}")
            if not existing:
                st.error("None of the selected columns exist in the current data.")
                return df
            return df[existing]
        except Exception as e:
            st.error(f"Error selecting columns: {e}")
            return df

    @staticmethod
    def pivot_table(df, index, columns, values, aggfunc):
        try:
            result = pd.pivot_table(df, values=values, index=index,
                                     columns=columns, aggfunc=aggfunc)
            return result.reset_index()
        except Exception as e:
            st.error(f"Error creating pivot: {e}")
            return df

    @staticmethod
    def create_relationship(df1, df2, on, how='inner'):
        try:
            return pd.merge(df1, df2, on=on, how=how)
        except Exception as e:
            st.error(f"Error creating relationship: {e}")
            return None

    @staticmethod
    def vlookup(df, lookup_col, lookup_df, lookup_on, return_col):
        try:
            mapping = dict(zip(lookup_df[lookup_on], lookup_df[return_col]))
            return df[lookup_col].map(mapping)
        except Exception as e:
            st.error(f"Error in VLOOKUP: {e}")
            return None

    @staticmethod
    def copy_column_into(df, ref_df, ref_column, new_column_name, align_by='position'):
        try:
            if ref_column not in ref_df.columns:
                st.error(f"Column '{ref_column}' not found in the reference source")
                return df
            df = df.copy()
            source_series = ref_df[ref_column]
            if align_by == 'position':
                values = source_series.reset_index(drop=True)
                values.index = df.reset_index(drop=True).index[:len(values)]
                new_col = pd.Series(index=df.index, dtype=object)
                n = min(len(df), len(source_series))
                new_col.iloc[:n] = source_series.iloc[:n].values
                df[new_column_name] = new_col
            else:
                df[new_column_name] = source_series.reindex(df.index)
            return df
        except Exception as e:
            st.error(f"Error copying column: {e}")
            return df

    @staticmethod
    def filter_by_reference(df, column, reference_df, ref_column, operation='in'):
        try:
            ref_values = reference_df[ref_column].unique()
            if operation == 'in':
                return df[df[column].isin(ref_values)]
            elif operation == 'not in':
                return df[~df[column].isin(ref_values)]
            return df
        except Exception as e:
            st.error(f"Error filtering by reference: {e}")
            return df

    @staticmethod
    def append_data(df1, df2):
        try:
            return pd.concat([df1, df2], ignore_index=True)
        except Exception as e:
            st.error(f"Error appending data: {e}")
            return df1

    @staticmethod
    def summarize_by_reference(df, group_col, agg_col, ref_df, ref_col, agg_func='sum'):
        try:
            agg = df.groupby(group_col)[agg_col].agg(agg_func).reset_index()
            return pd.merge(agg, ref_df, left_on=group_col, right_on=ref_col, how='left')
        except Exception as e:
            st.error(f"Error summarizing by reference: {e}")
            return df

    @staticmethod
    def build_template_output(df, template_headers, column_configs, all_sources=None):
        if len(df) == 0:
            st.warning("⚠️ Building template output, but the input data already has 0 rows "
                       "(an earlier step - likely a filter - removed everything). "
                       "The output will also have 0 rows.")
        result = pd.DataFrame(index=df.index)
        for header in template_headers:
            cfg = column_configs.get(header, {'mode': 'static', 'value': ''})
            mode = cfg.get('mode')
            try:
                if mode == 'direct':
                    col = cfg.get('source_column')
                    result[header] = df[col] if col in df.columns else np.nan

                elif mode == 'lookup':
                    ref_name = cfg.get('ref_source')
                    ref_df = (all_sources or {}).get(ref_name)
                    if ref_df is None:
                        st.error(f"Reference source '{ref_name}' not found for column '{header}'")
                        result[header] = np.nan
                        continue
                    key_col = cfg.get('key_column')
                    ref_key_col = cfg.get('ref_key_column')
                    ref_return_col = cfg.get('ref_return_column')
                    mapping = dict(zip(ref_df[ref_key_col], ref_df[ref_return_col]))
                    mapped = df[key_col].map(mapping)
                    default = cfg.get('default')
                    if default not in (None, ''):
                        mapped = mapped.fillna(default)
                    result[header] = mapped

                elif mode == 'formula':
                    result[header] = MultiSheetProcessor.evaluate_formula(df, cfg.get('formula', ''), all_sources)

                elif mode == 'static':
                    result[header] = cfg.get('value', '')

                else:
                    result[header] = np.nan

                result[header] = apply_output_format(result[header], cfg.get('format'))
            except Exception as e:
                st.error(f"Error building column '{header}': {e}")
                result[header] = np.nan

        return result.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Operation execution
# ---------------------------------------------------------------------------
SOURCE_REF_KEYS = ['source', 'lookup_source', 'ref_source', 'append_source']


def _warn_if_wiped(before_len, result, op_desc):
    if before_len > 0 and result is not None and len(result) == 0:
        st.warning(f"⚠️ Step wiped all rows to 0: {op_desc}. Check the column name, "
                   f"value, and data type used in this step.")
    return result


def apply_operation(df, operation, all_sources=None):
    op_type = operation['type']
    before_len = len(df)

    if op_type == 'filter':
        result = MultiSheetProcessor.filter_rows(df, operation['column'], operation['operation'], operation['value'])
        return _warn_if_wiped(before_len, result, describe_operation(operation))

    elif op_type == 'filter_multiple':
        result = MultiSheetProcessor.filter_multiple_rows(df, operation['conditions'], operation.get('logic', 'AND'))
        return _warn_if_wiped(before_len, result, describe_operation(operation))

    elif op_type == 'calc_column':
        return MultiSheetProcessor.add_calculated_column(df, operation['new_column'], operation['formula'], all_sources)

    elif op_type == 'build_template':
        return MultiSheetProcessor.build_template_output(
            df, operation['template_headers'], operation['column_configs'], all_sources
        )

    elif op_type == 'group_by':
        return MultiSheetProcessor.group_by(df, operation['group_column'], operation['aggregate_column'], operation['function'])

    elif op_type == 'sort':
        return MultiSheetProcessor.sort_data(df, operation['column'], operation['ascending'])

    elif op_type == 'rename':
        return MultiSheetProcessor.rename_column(df, operation['old_name'], operation['new_name'])

    elif op_type == 'drop':
        return MultiSheetProcessor.drop_column(df, operation['column'])

    elif op_type == 'select_columns':
        return MultiSheetProcessor.select_columns(df, operation['columns'])

    elif op_type == 'remove_duplicates':
        result = MultiSheetProcessor.remove_duplicates(df, operation.get('subset'), operation.get('keep', 'first'))
        return _warn_if_wiped(before_len, result, describe_operation(operation))

    elif op_type == 'handle_missing':
        return MultiSheetProcessor.handle_missing_values(
            df, operation.get('column'), operation['method'], operation.get('value')
        )

    elif op_type == 'change_dtype':
        return MultiSheetProcessor.change_data_type(df, operation['column'], operation['target_type'])

    elif op_type == 'find_replace':
        return MultiSheetProcessor.find_replace(
            df, operation.get('column'), operation['find_text'], operation['replace_text'],
            operation.get('match_case', False), operation.get('whole_cell', False)
        )

    elif op_type == 'split_column':
        return MultiSheetProcessor.split_column(
            df, operation['column'], operation['delimiter'], operation.get('new_names')
        )

    elif op_type == 'merge':
        if all_sources and operation['source'] in all_sources:
            return MultiSheetProcessor.create_relationship(
                df, all_sources[operation['source']], operation['on'], operation['how']
            )
        return df

    elif op_type == 'copy_column':
        if all_sources and operation['ref_source'] in all_sources:
            return MultiSheetProcessor.copy_column_into(
                df, all_sources[operation['ref_source']], operation['ref_column'], operation['new_column']
            )
        return df

    elif op_type == 'vlookup':
        if all_sources and operation['lookup_source'] in all_sources:
            result = MultiSheetProcessor.vlookup(
                df, operation['lookup_col'], all_sources[operation['lookup_source']],
                operation['lookup_on'], operation['return_col']
            )
            if result is not None:
                df = df.copy()
                df[operation['new_column']] = result
            return df
        return df

    elif op_type == 'filter_by_ref':
        if all_sources and operation['ref_source'] in all_sources:
            return MultiSheetProcessor.filter_by_reference(
                df, operation['column'], all_sources[operation['ref_source']],
                operation['ref_column'], operation.get('operation', 'in')
            )
        return df

    elif op_type == 'append':
        if all_sources and operation['append_source'] in all_sources:
            return MultiSheetProcessor.append_data(df, all_sources[operation['append_source']])
        return df

    elif op_type == 'summarize_by_ref':
        if all_sources and operation['ref_source'] in all_sources:
            return MultiSheetProcessor.summarize_by_reference(
                df, operation['group_col'], operation['agg_col'],
                all_sources[operation['ref_source']], operation['ref_col'],
                operation.get('agg_func', 'sum')
            )
        return df

    else:
        return df


def execute_all_operations(df, operations, all_sources):
    for op in operations:
        df = apply_operation(df, op, all_sources)
    return df


def extract_referenced_sources(recipe):
    refs = set()
    if recipe.get('primary_source'):
        refs.add(recipe['primary_source'])
    for op in recipe.get('operations', []):
        for key in SOURCE_REF_KEYS:
            if key in op and op[key]:
                refs.add(op[key])
        if op['type'] == 'calc_column' and op.get('referenced_sources'):
            refs.update(op['referenced_sources'])
        if op['type'] == 'build_template':
            for cfg in op.get('column_configs', {}).values():
                if cfg.get('mode') == 'lookup' and cfg.get('ref_source'):
                    refs.add(cfg['ref_source'])
                if cfg.get('mode') == 'formula' and cfg.get('referenced_sources'):
                    refs.update(cfg['referenced_sources'])
    return sorted(refs)


def _remap_formula_text(formula, referenced_sources, mapping):
    for old_name in referenced_sources:
        new_name = mapping.get(old_name)
        if new_name:
            formula = formula.replace(f"'{old_name}'", f"'{new_name}'").replace(f'"{old_name}"', f'"{new_name}"')
    return formula


def remap_operations(operations, mapping):
    new_ops = []
    for op in operations:
        new_op = dict(op)

        for key in SOURCE_REF_KEYS:
            if key in new_op and new_op[key] in mapping:
                new_op[key] = mapping[new_op[key]]

        if new_op['type'] == 'calc_column' and new_op.get('referenced_sources'):
            new_op['formula'] = _remap_formula_text(new_op['formula'], new_op['referenced_sources'], mapping)
            new_op['referenced_sources'] = [mapping.get(s, s) for s in new_op['referenced_sources']]

        if new_op['type'] == 'build_template':
            new_configs = {}
            for col, cfg in new_op.get('column_configs', {}).items():
                cfg = dict(cfg)
                if cfg.get('mode') == 'lookup' and cfg.get('ref_source') in mapping:
                    cfg['ref_source'] = mapping[cfg['ref_source']]
                if cfg.get('mode') == 'formula' and cfg.get('referenced_sources'):
                    cfg['formula'] = _remap_formula_text(cfg['formula'], cfg['referenced_sources'], mapping)
                    cfg['referenced_sources'] = [mapping.get(s, s) for s in cfg['referenced_sources']]
                new_configs[col] = cfg
            new_op['column_configs'] = new_configs

        new_ops.append(new_op)
    return new_ops


def build_recipe():
    return {
        'primary_source': st.session_state.active_source,
        'operations': st.session_state.operations,
        'timestamp': datetime.now().isoformat(),
    }


# ---------------------------------------------------------------------------
# Recipe management
# ---------------------------------------------------------------------------
def recompute_active_source():
    name = st.session_state.active_source
    if not name:
        return
    snapshot = st.session_state.source_snapshots.get(name)
    if snapshot is None:
        return
    result = snapshot.copy()
    for op in st.session_state.operations:
        result = apply_operation(result, op, st.session_state.data_sources)
    st.session_state.data_sources[name] = result


def push_undo_snapshot():
    st.session_state.operations_undo_stack.append(list(st.session_state.operations))
    if len(st.session_state.operations_undo_stack) > 25:
        st.session_state.operations_undo_stack.pop(0)


def undo_last_operation():
    if st.session_state.operations_undo_stack:
        st.session_state.operations = st.session_state.operations_undo_stack.pop()
        recompute_active_source()
        clear_step_editing_state()


def clear_step_editing_state():
    for key in list(st.session_state.keys()):
        if key.startswith("editing_step_"):
            st.session_state[key] = False


def describe_operation(op):
    t = op.get('type')
    try:
        if t == 'filter':
            return f"Filter: {op['column']} {op['operation']} {op['value']}"
        if t == 'filter_multiple':
            cond_count = len(op.get('conditions', []))
            logic = op.get('logic', 'AND')
            return f"Multiple Filters ({logic}): {cond_count} condition(s)"
        if t == 'calc_column':
            return f"Add column '{op['new_column']}' = {op['formula']}"
        if t == 'group_by':
            return f"Group by {op['group_column']}, {op['function']}({op['aggregate_column']})"
        if t == 'sort':
            return f"Sort by {op['column']} ({'asc' if op['ascending'] else 'desc'})"
        if t == 'rename':
            return f"Rename '{op['old_name']}' → '{op['new_name']}'"
        if t == 'drop':
            return f"Drop column '{op['column']}'"
        if t == 'select_columns':
            cols = op.get('columns', [])
            preview = ', '.join(cols[:4]) + (', …' if len(cols) > 4 else '')
            return f"Select {len(cols)} column(s) to keep: {preview}"
        if t == 'remove_duplicates':
            subset = op.get('subset')
            scope = f"columns {subset}" if subset else "all columns"
            keep_label = {'first': 'keep first', 'last': 'keep last', 'drop_all': 'drop all copies'}.get(op.get('keep', 'first'), 'keep first')
            return f"Remove duplicate rows (based on {scope}, {keep_label})"
        if t == 'handle_missing':
            col = op.get('column') or 'all columns'
            method_label = {
                'drop_rows': 'drop rows with missing values', 'fill_static': f"fill with '{op.get('value')}'",
                'fill_mean': 'fill with mean', 'fill_median': 'fill with median',
                'fill_mode': 'fill with most common value', 'fill_forward': 'fill forward',
                'fill_backward': 'fill backward',
            }.get(op.get('method'), op.get('method'))
            return f"Handle missing values in {col}: {method_label}"
        if t == 'change_dtype':
            return f"Change data type of '{op['column']}' to {op['target_type']}"
        if t == 'find_replace':
            col = op.get('column') or 'all text columns'
            return f"Find & Replace in {col}: '{op['find_text']}' → '{op['replace_text']}'"
        if t == 'split_column':
            return f"Split column '{op['column']}' by '{op['delimiter']}'"
        if t == 'merge':
            return f"Merge with '{op['source']}' on {op['on']} ({op['how']})"
        if t == 'vlookup':
            return f"VLOOKUP '{op['lookup_col']}' from '{op['lookup_source']}' → '{op['new_column']}'"
        if t == 'copy_column':
            return f"Copy column '{op['ref_column']}' from '{op['ref_source']}' → '{op['new_column']}'"
        if t == 'filter_by_ref':
            return f"Filter '{op['column']}' {op['operation']} ref '{op['ref_source']}'.{op['ref_column']}"
        if t == 'append':
            return f"Append rows from '{op['append_source']}'"
        if t == 'summarize_by_ref':
            return f"Summarize {op.get('agg_func', 'sum')}({op['agg_col']}) by {op['group_col']}, joined with '{op['ref_source']}'"
        if t == 'build_template':
            return f"Build output from template ({len(op.get('template_headers', []))} column(s))"
        if t == 'pivot':
            return f"Pivot: index={op['index']}, columns={op['columns']}, values={op['values']} ({op['aggfunc']})"
    except Exception:
        pass
    return t


def render_operation_edit_form(op, idx, baseline_df, all_sources):
    t = op.get('type')
    new_op = dict(op)
    cols = list(baseline_df.columns) if baseline_df is not None else []
    src_names = list(all_sources.keys())

    if t == 'filter':
        c1, c2 = st.columns(2)
        with c1:
            new_op['column'] = st.text_input("Column", op['column'], key=f"edit_col_{idx}")
        with c2:
            ops_list = ['==', '!=', '>', '<', '>=', '<=', 'contains']
            new_op['operation'] = st.selectbox("Operation", ops_list,
                                                index=ops_list.index(op['operation']) if op['operation'] in ops_list else 0,
                                                key=f"edit_op_{idx}")
        new_op['value'] = st.text_input("Value", op['value'], key=f"edit_val_{idx}")

    elif t == 'filter_multiple':
        st.info("Edit each condition individually")
        logic_opts = ['AND', 'OR']
        cur_logic = op.get('logic', 'AND')
        new_op['logic'] = st.radio(
            "Combine conditions using", logic_opts,
            index=logic_opts.index(cur_logic) if cur_logic in logic_opts else 0,
            horizontal=True, key=f"edit_mf_logic_{idx}",
            help="AND: rows must match every condition. OR: rows must match at least one condition."
        )
        new_conditions = []
        for i, cond in enumerate(op.get('conditions', [])):
            st.write(f"Condition {i+1}:")
            col1, col2, col3 = st.columns(3)
            with col1:
                cond_col = st.text_input("Column", cond.get('column', ''), key=f"edit_mf_col_{idx}_{i}")
            with col2:
                ops_list = ['==', '!=', '>', '<', '>=', '<=', 'contains']
                cond_op = st.selectbox("Operation", ops_list,
                                       index=ops_list.index(cond.get('operation', '==')) if cond.get('operation') in ops_list else 0,
                                       key=f"edit_mf_op_{idx}_{i}")
            with col3:
                cond_val = st.text_input("Value", cond.get('value', ''), key=f"edit_mf_val_{idx}_{i}")
            new_conditions.append({'column': cond_col, 'operation': cond_op, 'value': cond_val})
        new_op['conditions'] = new_conditions

    elif t == 'calc_column':
        new_op['new_column'] = st.text_input("New column name", op['new_column'], key=f"edit_newcol_{idx}")
        new_op['formula'] = st.text_area("Formula", op['formula'], key=f"edit_formula_{idx}")
        new_op['referenced_sources'] = detect_referenced_source_names(new_op['formula'], all_sources.keys())

    elif t == 'sort':
        new_op['column'] = st.text_input("Sort column", op['column'], key=f"edit_sortcol_{idx}")
        order = st.radio("Order", ["Ascending", "Descending"], index=0 if op['ascending'] else 1,
                          horizontal=True, key=f"edit_sortorder_{idx}")
        new_op['ascending'] = order == "Ascending"

    elif t == 'rename':
        new_op['old_name'] = st.text_input("Old name", op['old_name'], key=f"edit_oldname_{idx}")
        new_op['new_name'] = st.text_input("New name", op['new_name'], key=f"edit_newname_{idx}")

    elif t == 'drop':
        new_op['column'] = st.text_input("Column to drop", op['column'], key=f"edit_dropcol_{idx}")

    elif t == 'select_columns':
        options = cols if cols else op.get('columns', [])
        current = [c for c in op.get('columns', []) if c in options] or op.get('columns', [])
        new_op['columns'] = st.multiselect(
            "Columns to keep (order = output order)", options,
            default=current, key=f"edit_selectcols_{idx}"
        )

    elif t == 'group_by':
        new_op['group_column'] = st.text_input("Group column", op['group_column'], key=f"edit_groupcol_{idx}")
        new_op['aggregate_column'] = st.text_input("Aggregate column", op['aggregate_column'], key=f"edit_aggcol_{idx}")
        funcs = ['sum', 'mean', 'median', 'min', 'max', 'count', 'std']
        new_op['function'] = st.selectbox("Function", funcs,
                                           index=funcs.index(op['function']) if op['function'] in funcs else 0,
                                           key=f"edit_aggfunc_{idx}")

    elif t == 'vlookup':
        new_op['lookup_source'] = st.selectbox("Reference source", src_names,
                                                index=src_names.index(op['lookup_source']) if op['lookup_source'] in src_names else 0,
                                                key=f"edit_vlsrc_{idx}")
        new_op['lookup_col'] = st.text_input("Lookup column (current)", op['lookup_col'], key=f"edit_vlcol_{idx}")
        new_op['lookup_on'] = st.text_input("Lookup column (reference)", op['lookup_on'], key=f"edit_vlon_{idx}")
        new_op['return_col'] = st.text_input("Return column", op['return_col'], key=f"edit_vlret_{idx}")
        new_op['new_column'] = st.text_input("New column name", op['new_column'], key=f"edit_vlnew_{idx}")

    elif t == 'copy_column':
        new_op['ref_source'] = st.selectbox("Reference source", src_names,
                                             index=src_names.index(op['ref_source']) if op['ref_source'] in src_names else 0,
                                             key=f"edit_ccsrc_{idx}")
        new_op['ref_column'] = st.text_input("Column to copy", op['ref_column'], key=f"edit_cccol_{idx}")
        new_op['new_column'] = st.text_input("New column name", op['new_column'], key=f"edit_ccnew_{idx}")

    elif t == 'filter_by_ref':
        new_op['ref_source'] = st.selectbox("Reference source", src_names,
                                             index=src_names.index(op['ref_source']) if op['ref_source'] in src_names else 0,
                                             key=f"edit_fbrsrc_{idx}")
        new_op['column'] = st.text_input("Column to filter", op['column'], key=f"edit_fbrcol_{idx}")
        new_op['ref_column'] = st.text_input("Reference column", op['ref_column'], key=f"edit_fbrrefcol_{idx}")
        new_op['operation'] = st.selectbox("Operation", ['in', 'not in'],
                                            index=0 if op['operation'] == 'in' else 1, key=f"edit_fbrop_{idx}")

    elif t == 'append':
        new_op['append_source'] = st.selectbox("Source to append", src_names,
                                                index=src_names.index(op['append_source']) if op['append_source'] in src_names else 0,
                                                key=f"edit_appendsrc_{idx}")

    elif t == 'summarize_by_ref':
        new_op['ref_source'] = st.selectbox("Reference source", src_names,
                                             index=src_names.index(op['ref_source']) if op['ref_source'] in src_names else 0,
                                             key=f"edit_sumref_{idx}")
        new_op['group_col'] = st.text_input("Group column", op['group_col'], key=f"edit_sumgroup_{idx}")
        new_op['agg_col'] = st.text_input("Aggregate column", op['agg_col'], key=f"edit_sumagg_{idx}")
        new_op['ref_col'] = st.text_input("Reference join column", op['ref_col'], key=f"edit_sumrefcol_{idx}")
        funcs = ['sum', 'mean', 'count', 'min', 'max']
        cur_func = op.get('agg_func', 'sum')
        new_op['agg_func'] = st.selectbox("Function", funcs,
                                           index=funcs.index(cur_func) if cur_func in funcs else 0,
                                           key=f"edit_sumfunc_{idx}")

    elif t == 'pivot':
        new_op['index'] = st.text_input("Index column", op['index'], key=f"edit_pvindex_{idx}")
        new_op['columns'] = st.text_input("Columns", op['columns'], key=f"edit_pvcols_{idx}")
        new_op['values'] = st.text_input("Values", op['values'], key=f"edit_pvvals_{idx}")
        funcs = ['sum', 'mean', 'count', 'min', 'max']
        new_op['aggfunc'] = st.selectbox("Aggregation", funcs,
                                          index=funcs.index(op['aggfunc']) if op['aggfunc'] in funcs else 0,
                                          key=f"edit_pvagg_{idx}")

    elif t == 'merge':
        new_op['source'] = st.selectbox("Merge with", src_names,
                                         index=src_names.index(op['source']) if op['source'] in src_names else 0,
                                         key=f"edit_mergesrc_{idx}")
        new_op['on'] = st.text_input("Join column", op['on'], key=f"edit_mergeon_{idx}")
        how_opts = ['inner', 'left', 'right', 'outer']
        new_op['how'] = st.selectbox("Join type", how_opts,
                                      index=how_opts.index(op['how']) if op['how'] in how_opts else 0,
                                      key=f"edit_mergehow_{idx}")

    else:
        st.info("This step type doesn't support inline field editing yet — you can still reorder or delete it.")

    return new_op


_EXCEL_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _safe_excel_sheet_name(name, used_names):
    clean = _EXCEL_INVALID_SHEET_CHARS.sub("_", str(name)).strip() or "Sheet"
    base = clean[:31]
    candidate = base
    n = 1
    while candidate in used_names:
        suffix = f"_{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used_names.add(candidate)
    return candidate


def to_excel_bytes(dfs_by_name):
    output = io.BytesIO()
    used_names = set()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name, data in dfs_by_name.items():
            sheet_name = _safe_excel_sheet_name(name, used_names)
            data.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


def build_original_file_bytes(original_file_name, all_source_metadata, all_source_data):
    original = st.session_state.original_files.get(original_file_name)
    if original is None:
        return None, None

    ext = original['ext']
    raw_bytes = original['bytes']

    sources_by_sheet = {}
    for name, meta in all_source_metadata.items():
        if meta.get('file') == original_file_name and name in all_source_data:
            sources_by_sheet.setdefault(meta.get('sheet'), name)

    if ext == '.csv':
        source_name = sources_by_sheet.get('Sheet1')
        if source_name:
            return all_source_data[source_name].to_csv(index=False).encode('utf-8'), '.csv'
        return raw_bytes, '.csv'

    xl = pd.ExcelFile(io.BytesIO(raw_bytes))
    output = io.BytesIO()
    used_names = set()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name in xl.sheet_names:
            source_name = sources_by_sheet.get(sheet_name)
            if source_name:
                sheet_df = all_source_data[source_name]
            else:
                try:
                    sheet_df = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=sheet_name)
                except Exception:
                    sheet_df = pd.DataFrame()
            safe_name = _safe_excel_sheet_name(sheet_name, used_names)
            sheet_df.to_excel(writer, index=False, sheet_name=safe_name)
    return output.getvalue(), '.xlsx'


def build_styled_template_excel(template_bytes, sheet_name, header_row, headers, result_df):
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    header_excel_row = header_row + 1
    data_start_row = header_excel_row + 1

    header_col_map = {}
    for cell in ws[header_excel_row]:
        if cell.value is not None:
            header_col_map[str(cell.value).strip()] = cell.column

    sample_row_has_data = any(
        ws.cell(row=data_start_row, column=c).value is not None for c in header_col_map.values()
    )
    style_source_row = data_start_row if sample_row_has_data else header_excel_row

    col_styles = {}
    for header, col_idx in header_col_map.items():
        src_cell = ws.cell(row=style_source_row, column=col_idx)
        col_styles[header] = {
            "font": _copy_style(src_cell.font),
            "fill": _copy_style(src_cell.fill),
            "border": _copy_style(src_cell.border),
            "alignment": _copy_style(src_cell.alignment),
            "number_format": src_cell.number_format,
        }

    max_existing_row = ws.max_row
    if max_existing_row >= data_start_row:
        for row in ws.iter_rows(min_row=data_start_row, max_row=max_existing_row):
            for cell in row:
                cell.value = None

    for row_offset, (_, row_data) in enumerate(result_df.iterrows()):
        excel_row = data_start_row + row_offset
        for header in headers:
            col_idx = header_col_map.get(header)
            if col_idx is None:
                continue
            cell = ws.cell(row=excel_row, column=col_idx, value=row_data.get(header))
            style = col_styles.get(header)
            if style:
                cell.font = style["font"]
                cell.fill = style["fill"]
                cell.border = style["border"]
                cell.alignment = style["alignment"]
                if style["number_format"]:
                    cell.number_format = style["number_format"]

    last_row = data_start_row + len(result_df) - 1
    if header_col_map:
        min_col = min(header_col_map.values())
        max_col = max(header_col_map.values())
        try:
            for tbl in list(ws.tables.values()):
                tbl_range = tbl.ref
                if f"{header_excel_row}" in tbl_range.split(":")[0] or tbl.headerRowCount:
                    new_ref = (
                        f"{get_column_letter(min_col)}{header_excel_row}:"
                        f"{get_column_letter(max_col)}{max(last_row, header_excel_row)}"
                    )
                    tbl.ref = new_ref
        except Exception:
            pass

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ========================= MAIN APP =========================

st.markdown('<h1 class="main-header">📚 Multi-Sheet Data Processing Wizard</h1>', unsafe_allow_html=True)
st.markdown('<p class="main-subheader">Build your process once, save it, and reuse it on every new report</p>', unsafe_allow_html=True)

# Active navbar style overrides
current_p = st.session_state.current_page
active_key = "nav_sources_btn" if current_p == "Workspace" else ("nav_automation_btn" if current_p == "Automation" else ("nav_batch_btn" if current_p == "Batch Automation" else "nav_manual_btn"))
st.markdown(f"""
    <style>
    div.stButton > button[key="{active_key}"] {{
        background: linear-gradient(135deg, #00f2fe 0%, #4facfe 50%, #9b51e0 100%) !important;
        color: #ffffff !important;
        font-weight: 700 !important;
        border: none !important;
        box-shadow: 0 0 20px rgba(0, 242, 254, 0.4) !important;
    }}
    </style>
""", unsafe_allow_html=True)

# Top Navigation Bar
st.markdown('<div class="navbar-container">', unsafe_allow_html=True)
nav_col1, nav_col2, nav_col3, nav_col4 = st.columns([1, 1, 1, 1])
with nav_col1:
    if st.button("📁 Data Sources", key="nav_sources_btn", use_container_width=True):
        st.session_state.current_page = "Workspace"
        st.rerun()
with nav_col2:
    if st.button("🤖 Automation", key="nav_automation_btn", use_container_width=True):
        st.session_state.current_page = "Automation"
        st.rerun()
with nav_col3:
    if st.button("📦 Batch Automation", key="nav_batch_btn", use_container_width=True):
        st.session_state.current_page = "Batch Automation"
        st.rerun()
with nav_col4:
    if st.button("📖 User Manual", key="nav_manual_btn", use_container_width=True):
        st.session_state.current_page = "Manual"
        st.rerun()
st.markdown('</div>', unsafe_allow_html=True)

# Navigation Notification Banner
st.markdown(
    '<div style="text-align: center; margin-bottom: 15px; font-size: 0.82rem; color: #64748b; font-family: \'Inter\', sans-serif;">'
    '💡 <strong>Tip:</strong> Use the top navbar to switch views. Load reference sheets under Data Sources first before running lookups.'
    '</div>',
    unsafe_allow_html=True
)

page = st.session_state.current_page

if page == "Workspace":
    if not st.session_state.data_sources:
        # Show clean browse/upload glass card
        st.markdown('<div class="glass-card">', unsafe_allow_html=True)
        st.markdown('<h2 style="font-family: \'Outfit\', sans-serif; color: #00f2fe; margin-top: 0;">📁 Load Data Sources</h2>', unsafe_allow_html=True)
        st.markdown('<p style="font-size: 1rem; color: #cbd5e1; margin-bottom: 25px;">Upload your CSV or Excel files to begin processing your sheets.</p>', unsafe_allow_html=True)

        uploaded_files = st.file_uploader(
            "Upload files (CSV or Excel)",
            type=['csv', 'xlsx', 'xls'],
            accept_multiple_files=True,
            help="Upload multiple files at once. Excel files will load all sheets.",
            key="workspace_empty_uploader"
        )

        if uploaded_files:
            if st.button("🔄 Load All Files", use_container_width=True, key="load_all_files_btn"):
                newly_loaded, overwritten, failed = 0, [], []
                pending_passwords = []
                pending_header_review = []
                
                # First pass: check for password-protected files and header ambiguity
                for file in uploaded_files:
                    ext = Path(file.name).suffix.lower()
                    file_bytes = file.getvalue()
                    
                    # Check for password protection
                    if ext in ('.xlsx', '.xls'):
                        if is_excel_password_protected(file_bytes):
                            pending_passwords.append({
                                'file': file,
                                'file_bytes': file_bytes,
                                'name': file.name,
                                'ext': ext
                            })
                            continue
                    
                    # Check for header ambiguity in each sheet
                    try:
                        if ext == ".csv":
                            raw_preview = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=20, dtype=str)
                            ambiguous, suggested_row = analyze_header_ambiguity(raw_preview)
                            if ambiguous:
                                pending_header_review.append({
                                    'file_name': file.name,
                                    'file_stem': Path(file.name).stem,
                                    'ext': ext,
                                    'file_bytes': file_bytes,
                                    'sheet_name': None,
                                    'sheet_label': 'Sheet1',
                                    'raw_preview': raw_preview,
                                    'suggested_row': suggested_row,
                                    'multi_sheet': False
                                })
                        else:
                            xl = pd.ExcelFile(io.BytesIO(file_bytes))
                            for sheet_name in xl.sheet_names:
                                raw_preview = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, 
                                                            header=None, nrows=20, dtype=str)
                                ambiguous, suggested_row = analyze_header_ambiguity(raw_preview)
                                if ambiguous:
                                    pending_header_review.append({
                                        'file_name': file.name,
                                        'file_stem': Path(file.name).stem,
                                        'ext': ext,
                                        'file_bytes': file_bytes,
                                        'sheet_name': sheet_name,
                                        'sheet_label': sheet_name,
                                        'raw_preview': raw_preview,
                                        'suggested_row': suggested_row,
                                        'multi_sheet': len(xl.sheet_names) > 1
                                    })
                    except Exception as e:
                        failed.append(f"{file.name}: {str(e)}")
                
                # Handle password-protected files
                if pending_passwords:
                    st.session_state.pending_password_files = pending_passwords
                    st.markdown('<div class="password-box">', unsafe_allow_html=True)
                    st.markdown('<h4>🔐 Password-Protected Files Detected</h4>', unsafe_allow_html=True)
                    st.markdown(f"<p>{len(pending_passwords)} file(s) are password protected.</p>", unsafe_allow_html=True)
                    
                    apply_all = st.checkbox("Apply the same password to all locked files", value=True, key="apply_password_all")
                    
                    if apply_all:
                        master_password = st.text_input(
                            "Enter password for all locked files",
                            type="password",
                            key="master_password_input",
                            placeholder="Enter password to unlock all files..."
                        )
                        unlock_all_btn = st.button("🔓 Unlock All Files", use_container_width=True, key="unlock_all_btn")
                        
                        if unlock_all_btn and master_password:
                            for pwd_file in pending_passwords:
                                st.session_state.password_cache[pwd_file['name']] = master_password
                            st.rerun()
                    else:
                        st.caption("Enter passwords for each file individually:")
                        for pwd_file in pending_passwords:
                            col1, col2 = st.columns([3, 1])
                            with col1:
                                pwd = st.text_input(
                                    f"Password for {pwd_file['name']}",
                                    type="password",
                                    key=f"pwd_{pwd_file['name']}",
                                    placeholder="Enter password..."
                                )
                            with col2:
                                if st.button("🔓", key=f"unlock_{pwd_file['name']}", help="Unlock this file"):
                                    if pwd:
                                        st.session_state.password_cache[pwd_file['name']] = pwd
                                        st.rerun()
                    
                    st.markdown('</div>', unsafe_allow_html=True)
                
                # Handle header review
                if pending_header_review:
                    st.session_state.pending_header_review = pending_header_review
                    st.markdown('<div class="header-review-box">', unsafe_allow_html=True)
                    st.markdown('<h4>📋 Header Row Review Needed</h4>', unsafe_allow_html=True)
                    st.markdown(f"<p>{len(pending_header_review)} sheet(s) have ambiguous header rows. Please confirm the correct header row for each.</p>", unsafe_allow_html=True)
                    
                    for item in pending_header_review:
                        with st.expander(f"📄 {item['file_name']} — {item['sheet_label']}", expanded=True):
                            if item['raw_preview'] is not None and len(item['raw_preview']) > 0:
                                st.caption("Raw preview (no header applied yet) — row numbers start at 0:")
                                preview_display = item['raw_preview'].copy()
                                preview_display.index = [f"Row {i}" for i in range(len(preview_display))]
                                st.dataframe(preview_display, use_container_width=True)
                                max_row = len(item['raw_preview']) - 1
                            else:
                                st.caption("Couldn't generate a preview for this sheet — pick row 0 if unsure.")
                                max_row = 0
                            
                            st.number_input(
                                "Which row number is the header?",
                                min_value=0, max_value=max(max_row, 0),
                                value=min(item['suggested_row'], max(max_row, 0)),
                                key=f"header_row_choice_{item['file_name']}_{item['sheet_label']}",
                                help="Row 0 is the very first row of the file/sheet."
                            )
                            st.caption(f"💡 Suggested: Row {item['suggested_row']}")
                    
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        if st.button("✅ Confirm Headers & Load", use_container_width=True, key="confirm_header_review_btn"):
                            for item in pending_header_review:
                                chosen_row = st.session_state.get(
                                    f"header_row_choice_{item['file_name']}_{item['sheet_label']}", 
                                    item['suggested_row']
                                )
                                try:
                                    if item['ext'] == ".csv":
                                        df = MultiSheetProcessor.load_csv_with_header(item['file_bytes'], header_row=chosen_row)
                                        sheets = {"Sheet1": df}
                                    else:
                                        df = MultiSheetProcessor.load_excel_sheet_with_header(
                                            item['file_bytes'], item['sheet_name'], header_row=chosen_row
                                        )
                                        sheets = {item['sheet_label']: df}
                                    
                                    source_name = f"{item['file_stem']}_{item['sheet_label']}" if item['multi_sheet'] else item['file_stem']
                                    if source_name in st.session_state.data_sources:
                                        overwritten.append(source_name)
                                        if st.session_state.active_source == source_name:
                                            st.session_state.operations = []
                                            st.session_state.operations_undo_stack = []
                                    
                                    st.session_state.data_sources[source_name] = df
                                    st.session_state.source_snapshots[source_name] = df.copy()
                                    st.session_state.source_metadata[source_name] = {
                                        'file': item['file_name'],
                                        'sheet': item['sheet_label'],
                                        'rows': len(df),
                                        'columns': len(df.columns),
                                        'header_row': chosen_row
                                    }
                                    newly_loaded += 1
                                except Exception as e:
                                    failed.append(f"{item['file_name']} ({item['sheet_label']}): {e}")
                            
                            st.session_state.pending_header_review = []
                            st.rerun()
                    with rc2:
                        if st.button("🚫 Skip These Sheets", use_container_width=True, key="skip_header_review_btn"):
                            st.session_state.pending_header_review = []
                            st.info("Skipped sheets that needed header review.")
                            st.rerun()
                    
                    st.markdown('</div>', unsafe_allow_html=True)
                
                # Load files that don't need review or password
                for file in uploaded_files:
                    file_name = Path(file.name).stem
                    ext = Path(file.name).suffix.lower()
                    file_bytes = file.getvalue()
                    
                    # Skip if password protected (handled above)
                    if ext in ('.xlsx', '.xls') and is_excel_password_protected(file_bytes):
                        continue
                    
                    # Skip if in pending header review
                    skip = False
                    for item in pending_header_review:
                        if item['file_name'] == file.name:
                            skip = True
                            break
                    if skip:
                        continue
                    
                    # Check if we have a password for this file
                    password = st.session_state.password_cache.get(file.name)
                    
                    try:
                        if ext == ".csv":
                            # Load with header detection
                            raw_preview = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=20, dtype=str)
                            detected_row = detect_header_row(raw_preview)
                            df = pd.read_csv(io.BytesIO(file_bytes), header=detected_row)
                            df.columns = dedupe_column_names(df.columns)
                            for col in df.columns:
                                converted = try_convert_datetime_column(df[col])
                                if converted is not None:
                                    df[col] = converted
                            
                            st.session_state.original_files[file.name] = {'bytes': file_bytes, 'ext': ext}
                            source_name = file_name
                            if source_name in st.session_state.data_sources:
                                overwritten.append(source_name)
                                if st.session_state.active_source == source_name:
                                    st.session_state.operations = []
                                    st.session_state.operations_undo_stack = []
                            st.session_state.data_sources[source_name] = df
                            st.session_state.source_snapshots[source_name] = df.copy()
                            st.session_state.source_metadata[source_name] = {
                                'file': file.name, 'sheet': 'Sheet1',
                                'rows': len(df), 'columns': len(df.columns), 'header_row': detected_row
                            }
                            newly_loaded += 1
                            
                        elif ext in (".xlsx", ".xls"):
                            # Load all sheets with header detection
                            sheets, sheet_pending = MultiSheetProcessor.load_multi_sheet_file_with_header_detection(file_bytes, ext)
                            
                            if sheet_pending:
                                # Add to pending review
                                for item in sheet_pending:
                                    item['file_name'] = file.name
                                    item['file_stem'] = file_name
                                    item['file_bytes'] = file_bytes
                                    item['ext'] = ext
                                    item['multi_sheet'] = len(sheets) > 1 if sheets else False
                                st.session_state.pending_header_review.extend(sheet_pending)
                                continue
                            
                            if sheets:
                                st.session_state.original_files[file.name] = {'bytes': file_bytes, 'ext': ext}
                                
                                multi_sheet = len(sheets) > 1
                                for sheet_name, df in sheets.items():
                                    source_name = f"{file_name}_{sheet_name}" if multi_sheet else file_name
                                    if source_name in st.session_state.data_sources:
                                        overwritten.append(source_name)
                                        if st.session_state.active_source == source_name:
                                            st.session_state.operations = []
                                            st.session_state.operations_undo_stack = []
                                    st.session_state.data_sources[source_name] = df
                                    st.session_state.source_snapshots[source_name] = df.copy()
                                    st.session_state.source_metadata[source_name] = {
                                        'file': file.name, 'sheet': sheet_name,
                                        'rows': len(df), 'columns': len(df.columns)
                                    }
                                    newly_loaded += 1
                            else:
                                failed.append(f"{file.name}: No sheets loaded")
                    except Exception as e:
                        failed.append(f"{file.name}: {str(e)}")
                
                if newly_loaded:
                    st.success(f"✅ Loaded {newly_loaded} sheet(s). Total sources: {len(st.session_state.data_sources)}")
                if overwritten:
                    st.warning(f"⚠️ Replaced existing source(s) with new data: {', '.join(overwritten)}")
                if failed:
                    st.error(f"⚠️ Could not read: {', '.join(failed)}")
                st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

    else:
        # Split workspace column layout
        col_left, col_right = st.columns([1.2, 3])

        with col_left:
            st.markdown('<div class="glass-card" style="padding: 15px; margin-bottom: 15px;">', unsafe_allow_html=True)
            st.markdown('<h3 style="font-family: \'Outfit\', sans-serif; color: #00f2fe; margin: 0 0 10px 0;">📁 Data Sources</h3>', unsafe_allow_html=True)
            with st.expander("➕ Load More Files", expanded=False):
                uploaded_files = st.file_uploader(
                    "Upload more files",
                    type=['csv', 'xlsx', 'xls'],
                    accept_multiple_files=True,
                    key="add_more_uploader"
                )
                if uploaded_files:
                    if st.button("🔄 Load Files", use_container_width=True, key="load_more_files_btn"):
                        # Reuse the same loading logic (simplified version)
                        newly_loaded, overwritten, failed = 0, [], []
                        pending_passwords = []
                        pending_header_review = []
                        
                        for file in uploaded_files:
                            ext = Path(file.name).suffix.lower()
                            file_bytes = file.getvalue()
                            
                            if ext in ('.xlsx', '.xls') and is_excel_password_protected(file_bytes):
                                pending_passwords.append({'file': file, 'file_bytes': file_bytes, 'name': file.name, 'ext': ext})
                                continue
                            
                            try:
                                if ext == ".csv":
                                    raw_preview = pd.read_csv(io.BytesIO(file_bytes), header=None, nrows=20, dtype=str)
                                    ambiguous, suggested_row = analyze_header_ambiguity(raw_preview)
                                    if ambiguous:
                                        pending_header_review.append({
                                            'file_name': file.name, 'file_stem': Path(file.name).stem,
                                            'ext': ext, 'file_bytes': file_bytes,
                                            'sheet_name': None, 'sheet_label': 'Sheet1',
                                            'raw_preview': raw_preview, 'suggested_row': suggested_row,
                                            'multi_sheet': False
                                        })
                                    else:
                                        detected_row = detect_header_row(raw_preview)
                                        df = pd.read_csv(io.BytesIO(file_bytes), header=detected_row)
                                        df.columns = dedupe_column_names(df.columns)
                                        for col in df.columns:
                                            converted = try_convert_datetime_column(df[col])
                                            if converted is not None:
                                                df[col] = converted
                                        st.session_state.original_files[file.name] = {'bytes': file_bytes, 'ext': ext}
                                        source_name = Path(file.name).stem
                                        if source_name in st.session_state.data_sources:
                                            overwritten.append(source_name)
                                        st.session_state.data_sources[source_name] = df
                                        st.session_state.source_snapshots[source_name] = df.copy()
                                        st.session_state.source_metadata[source_name] = {
                                            'file': file.name, 'sheet': 'Sheet1',
                                            'rows': len(df), 'columns': len(df.columns), 'header_row': detected_row
                                        }
                                        newly_loaded += 1
                                else:
                                    sheets, sheet_pending = MultiSheetProcessor.load_multi_sheet_file_with_header_detection(file_bytes, ext)
                                    if sheet_pending:
                                        for item in sheet_pending:
                                            item['file_name'] = file.name
                                            item['file_stem'] = Path(file.name).stem
                                            item['file_bytes'] = file_bytes
                                            item['ext'] = ext
                                            item['multi_sheet'] = len(sheets) > 1 if sheets else False
                                        pending_header_review.extend(sheet_pending)
                                    elif sheets:
                                        st.session_state.original_files[file.name] = {'bytes': file_bytes, 'ext': ext}
                                        multi_sheet = len(sheets) > 1
                                        for sheet_name, df in sheets.items():
                                            source_name = f"{Path(file.name).stem}_{sheet_name}" if multi_sheet else Path(file.name).stem
                                            if source_name in st.session_state.data_sources:
                                                overwritten.append(source_name)
                                            st.session_state.data_sources[source_name] = df
                                            st.session_state.source_snapshots[source_name] = df.copy()
                                            st.session_state.source_metadata[source_name] = {
                                                'file': file.name, 'sheet': sheet_name,
                                                'rows': len(df), 'columns': len(df.columns)
                                            }
                                            newly_loaded += 1
                            except Exception as e:
                                failed.append(f"{file.name}: {str(e)}")
                        
                        if pending_header_review:
                            st.session_state.pending_header_review.extend(pending_header_review)
                        if pending_passwords:
                            st.session_state.pending_password_files.extend(pending_passwords)
                        
                        if newly_loaded:
                            st.success(f"✅ Loaded {newly_loaded} more sheet(s).")
                        if overwritten:
                            st.warning(f"⚠️ Replaced existing source(s): {', '.join(overwritten)}")
                        if failed:
                            st.error(f"⚠️ Could not read: {', '.join(failed)}")
                        st.rerun()

            # Show pending header review if any
            if st.session_state.pending_header_review:
                st.markdown('<div class="header-review-box">', unsafe_allow_html=True)
                st.markdown('<h4>📋 Header Row Review Needed</h4>', unsafe_allow_html=True)
                st.markdown(f"<p>{len(st.session_state.pending_header_review)} sheet(s) need header confirmation.</p>", unsafe_allow_html=True)
                
                for item in st.session_state.pending_header_review:
                    with st.expander(f"📄 {item['file_name']} — {item['sheet_label']}", expanded=True):
                        if item['raw_preview'] is not None and len(item['raw_preview']) > 0:
                            preview_display = item['raw_preview'].copy()
                            preview_display.index = [f"Row {i}" for i in range(len(preview_display))]
                            st.dataframe(preview_display, use_container_width=True)
                            max_row = len(item['raw_preview']) - 1
                        else:
                            max_row = 0
                        st.number_input(
                            "Which row number is the header?",
                            min_value=0, max_value=max(max_row, 0),
                            value=min(item['suggested_row'], max(max_row, 0)),
                            key=f"header_row_choice_{item['file_name']}_{item['sheet_label']}",
                            help="Row 0 is the very first row of the file/sheet."
                        )
                        st.caption(f"💡 Suggested: Row {item['suggested_row']}")
                
                rc1, rc2 = st.columns(2)
                with rc1:
                    if st.button("✅ Confirm Headers & Load", use_container_width=True, key="confirm_header_review_btn2"):
                        for item in st.session_state.pending_header_review:
                            chosen_row = st.session_state.get(
                                f"header_row_choice_{item['file_name']}_{item['sheet_label']}", 
                                item['suggested_row']
                            )
                            try:
                                if item['ext'] == ".csv":
                                    df = MultiSheetProcessor.load_csv_with_header(item['file_bytes'], header_row=chosen_row)
                                else:
                                    df = MultiSheetProcessor.load_excel_sheet_with_header(
                                        item['file_bytes'], item['sheet_name'], header_row=chosen_row
                                    )
                                
                                source_name = f"{item['file_stem']}_{item['sheet_label']}" if item['multi_sheet'] else item['file_stem']
                                if source_name in st.session_state.data_sources:
                                    if st.session_state.active_source == source_name:
                                        st.session_state.operations = []
                                        st.session_state.operations_undo_stack = []
                                st.session_state.data_sources[source_name] = df
                                st.session_state.source_snapshots[source_name] = df.copy()
                                st.session_state.source_metadata[source_name] = {
                                    'file': item['file_name'], 'sheet': item['sheet_label'],
                                    'rows': len(df), 'columns': len(df.columns), 'header_row': chosen_row
                                }
                            except Exception as e:
                                st.error(f"Error loading {item['file_name']} ({item['sheet_label']}): {e}")
                        
                        st.session_state.pending_header_review = []
                        st.rerun()
                with rc2:
                    if st.button("🚫 Skip These Sheets", use_container_width=True, key="skip_header_review_btn2"):
                        st.session_state.pending_header_review = []
                        st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

            # Show pending password files if any
            if st.session_state.pending_password_files:
                st.markdown('<div class="password-box">', unsafe_allow_html=True)
                st.markdown('<h4>🔐 Password-Protected Files</h4>', unsafe_allow_html=True)
                for pwd_file in st.session_state.pending_password_files:
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        pwd = st.text_input(
                            f"Password for {pwd_file['name']}",
                            type="password",
                            key=f"pwd_{pwd_file['name']}_2",
                            placeholder="Enter password..."
                        )
                    with col2:
                        if st.button("🔓", key=f"unlock_{pwd_file['name']}_2", help="Unlock this file"):
                            if pwd:
                                st.session_state.password_cache[pwd_file['name']] = pwd
                                # Try to load the file
                                try:
                                    if pwd_file['ext'] == ".csv":
                                        sheets, status = load_csv_with_password(pwd_file['file_bytes'], pwd)
                                    else:
                                        sheets, status = load_excel_with_password(pwd_file['file_bytes'], pwd)
                                    
                                    if status == "success" and sheets:
                                        for sheet_name, df in sheets.items():
                                            source_name = f"{Path(pwd_file['name']).stem}_{sheet_name}" if len(sheets) > 1 else Path(pwd_file['name']).stem
                                            st.session_state.data_sources[source_name] = df
                                            st.session_state.source_snapshots[source_name] = df.copy()
                                            st.session_state.source_metadata[source_name] = {
                                                'file': pwd_file['name'], 'sheet': sheet_name,
                                                'rows': len(df), 'columns': len(df.columns)
                                            }
                                        st.session_state.pending_password_files = [f for f in st.session_state.pending_password_files if f['name'] != pwd_file['name']]
                                        st.success(f"✅ Loaded {pwd_file['name']}")
                                        st.rerun()
                                    else:
                                        st.error(f"Failed to load {pwd_file['name']}: {status}")
                                        st.session_state.password_cache.pop(pwd_file['name'], None)
                                except Exception as e:
                                    st.error(f"Error loading {pwd_file['name']}: {str(e)}")
                                    st.session_state.password_cache.pop(pwd_file['name'], None)
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

            # Loaded sources list
            st.markdown('<div class="glass-card" style="padding: 15px; margin-bottom: 15px;">', unsafe_allow_html=True)
            st.markdown('<h4 style="margin: 0 0 10px 0; color: #cbd5e1; font-family: \'Outfit\', sans-serif;">📊 Loaded Sources</h4>', unsafe_allow_html=True)
            for source_name, df in list(st.session_state.data_sources.items()):
                col_btn, col_del = st.columns([4, 1])
                with col_btn:
                    is_active = st.session_state.active_source == source_name
                    btn_label = f"📊 {source_name}" + (" ✅" if is_active else "")
                    if st.button(btn_label, key=f"activate_{source_name}", use_container_width=True):
                        if st.session_state.active_source != source_name:
                            st.session_state.operations = []
                            st.session_state.operations_undo_stack = []
                            st.session_state.filter_conditions = []
                            clear_step_editing_state()
                        st.session_state.active_source = source_name
                        st.session_state.current_operation = None
                        if source_name not in st.session_state.source_snapshots:
                            st.session_state.source_snapshots[source_name] = st.session_state.data_sources[source_name].copy()
                        st.rerun()
                    meta = st.session_state.source_metadata.get(source_name, {})
                    st.caption(f"Rows: {meta.get('rows', 0)} | Cols: {meta.get('columns', 0)}")
                with col_del:
                    if st.button("❌", key=f"remove_{source_name}", help="Remove this source"):
                        del st.session_state.data_sources[source_name]
                        st.session_state.source_metadata.pop(source_name, None)
                        if st.session_state.active_source == source_name:
                            st.session_state.active_source = None
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

            # Relationships creator
            if len(st.session_state.data_sources) >= 2:
                st.markdown('<div class="glass-card" style="padding: 15px; margin-bottom: 15px;">', unsafe_allow_html=True)
                st.markdown('<h4 style="margin: 0 0 10px 0; color: #cbd5e1; font-family: \'Outfit\', sans-serif;">🔗 Relationships</h4>', unsafe_allow_html=True)
                with st.expander("Define Relationship"):
                    col1, col2 = st.columns(2)
                    with col1:
                        source1 = st.selectbox("Source 1", list(st.session_state.data_sources.keys()), key="rel_source1")
                    with col2:
                        source2 = st.selectbox("Source 2", list(st.session_state.data_sources.keys()), key="rel_source2")

                    if source1 != source2:
                        df1 = st.session_state.data_sources[source1]
                        df2 = st.session_state.data_sources[source2]
                        common_cols = [c for c in df1.columns if c in df2.columns]

                        if common_cols:
                            on_col = st.selectbox("Join on column", common_cols, key="rel_on_col")
                            how = st.selectbox("Join type", ['inner', 'left', 'right', 'outer'], key="rel_how")

                            if st.button("✅ Create Relationship", use_container_width=True, key="create_rel_btn"):
                                result = MultiSheetProcessor.create_relationship(df1, df2, on_col, how)
                                if result is not None:
                                    rel_name = f"{source1}_{source2}_joined"
                                    st.session_state.data_sources[rel_name] = result
                                    st.session_state.source_metadata[rel_name] = {
                                        'file': 'Relationship',
                                        'sheet': 'Joined',
                                        'rows': len(result),
                                        'columns': len(result.columns),
                                        'relationship': f'{source1} ↔ {source2} on {on_col}'
                                    }
                                    st.session_state.relationships.append({
                                        'source1': source1, 'source2': source2, 'on': on_col,
                                        'how': how, 'result': rel_name
                                    })
                                    st.success(f"✅ Created relationship: {rel_name}")
                                    st.caption(
                                        "Note: to include this join in a saved recipe, make the joined "
                                        f"source ('{rel_name}') your active source and continue building from there."
                                    )
                                    st.rerun()
                        else:
                            st.warning("No common columns found between these sources")
                st.markdown('</div>', unsafe_allow_html=True)

            # Recipe Timeline
            if st.session_state.operations:
                st.markdown('<div class="glass-card" style="padding: 15px; margin-bottom: 15px;">', unsafe_allow_html=True)
                st.markdown('<h4 style="margin: 0 0 10px 0; color: #cbd5e1; font-family: \'Outfit\', sans-serif;">📝 Recipe</h4>', unsafe_allow_html=True)
                
                if st.button("↩️ Undo Last Change", use_container_width=True, key="undo_btn",
                              disabled=(len(st.session_state.operations_undo_stack) == 0)):
                    undo_last_operation()
                    st.rerun()

                ops = st.session_state.operations
                baseline_df = st.session_state.source_snapshots.get(st.session_state.active_source)

                for i, op in enumerate(ops):
                    s1, s2, s3, s4, s5 = st.columns([5, 1, 1, 1, 1])
                    with s1:
                        st.markdown(f"""
                        <div class="timeline-step">
                            <div class="timeline-step-content">
                                <strong>Step {i+1}: {op['type'].upper()}</strong><br/>
                                <span style="font-size: 0.8rem; color: #cbd5e1;">{describe_operation(op)}</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    with s2:
                        if st.button("⬆️", key=f"up_{i}", disabled=(i == 0), help="Move step up"):
                            push_undo_snapshot()
                            ops[i-1], ops[i] = ops[i], ops[i-1]
                            recompute_active_source()
                            clear_step_editing_state()
                            st.rerun()
                    with s3:
                        if st.button("⬇️", key=f"down_{i}", disabled=(i == len(ops) - 1), help="Move step down"):
                            push_undo_snapshot()
                            ops[i+1], ops[i] = ops[i], ops[i+1]
                            recompute_active_source()
                            clear_step_editing_state()
                            st.rerun()
                    with s4:
                        if st.button("✏️", key=f"edit_toggle_{i}", help="Edit this step"):
                            st.session_state[f"editing_step_{i}"] = not st.session_state.get(f"editing_step_{i}", False)
                            st.rerun()
                    with s5:
                        if st.button("🗑️", key=f"delete_{i}", help="Delete this step"):
                            push_undo_snapshot()
                            ops.pop(i)
                            recompute_active_source()
                            clear_step_editing_state()
                            st.rerun()

                    if st.session_state.get(f"editing_step_{i}", False):
                        with st.expander(f"Editing step {i+1}", expanded=True):
                            edited_op = render_operation_edit_form(op, i, baseline_df, st.session_state.data_sources)
                            if st.button("✅ Save Changes", key=f"save_edit_{i}", use_container_width=True):
                                push_undo_snapshot()
                                ops[i] = edited_op
                                recompute_active_source()
                                st.session_state[f"editing_step_{i}"] = False
                                st.success(f"Step {i+1} updated")
                                st.rerun()

                recipe = build_recipe()
                recipe_json = json.dumps(recipe, indent=2)

                st.download_button(
                    "💾 Download Recipe (.json)",
                    data=recipe_json,
                    file_name=f"recipe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_recipe_btn"
                )

                with st.expander("View recipe JSON"):
                    st.code(recipe_json, language='json')

                if st.button("🔄 Reset All", use_container_width=True, key="reset_all_btn"):
                    st.session_state.data_sources = {}
                    st.session_state.source_metadata = {}
                    st.session_state.operations = []
                    st.session_state.active_source = None
                    st.session_state.relationships = []
                    st.session_state.current_operation = None
                    st.session_state.loaded_recipe = None
                    st.session_state.recipe_mapping = {}
                    st.session_state.source_snapshots = {}
                    st.session_state.operations_undo_stack = []
                    st.session_state.batch_results = {}
                    st.session_state.filter_conditions = []
                    st.session_state.template_headers = []
                    st.session_state.template_source_file = None
                    st.session_state.pending_header_review = []
                    st.session_state.original_files = {}
                    st.session_state.pending_password_files = []
                    st.session_state.password_cache = {}
                    for key in list(st.session_state.keys()):
                        if key.startswith("editing_step_") or key.startswith("tpl_") or key.startswith("header_row_choice_"):
                            del st.session_state[key]
                    st.success("✅ Reset complete!")
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        with col_right:
            if st.session_state.active_source:
                source_name = st.session_state.active_source
                df = st.session_state.data_sources[source_name]

                st.divider()
                meta = st.session_state.source_metadata.get(source_name, {})

                st.markdown('<div class="glass-card">', unsafe_allow_html=True)
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Active Source", source_name)
                with col2:
                    st.metric("Rows", len(df))
                with col3:
                    st.metric("Columns", len(df.columns))
                with col4:
                    st.metric("Operations", len(st.session_state.operations))
                st.markdown('</div>', unsafe_allow_html=True)

                with st.expander("🔍 Data Preview", expanded=False):
                    tab1, tab2, tab3 = st.tabs(["📊 Data", "📈 Statistics", "📋 Column Info"])
                    with tab1:
                        st.dataframe(df.head(100), use_container_width=True)
                    with tab2:
                        numeric_cols = df.select_dtypes(include=['number']).columns
                        if len(numeric_cols) > 0:
                            st.dataframe(df[numeric_cols].describe(), use_container_width=True)
                        else:
                            st.info("No numeric columns for statistics")
                    with tab3:
                        col_info = pd.DataFrame({
                            'Column': df.columns,
                            'Type': df.dtypes.astype(str),
                            'Null Count': df.isnull().sum().values,
                            'Unique Values': df.nunique().values
                        })
                        st.dataframe(col_info, use_container_width=True)

                if len(df.columns) == 0:
                    st.error(
                        "⚠️ The active source currently has no columns (an earlier step, like "
                        "Drop Column or Group By, may have removed them all). Undo the last "
                        "change in the Recipe panel, or pick a different data source."
                    )
                    st.stop()

                # Reference Operations
                if len(st.session_state.data_sources) > 1:
                    st.divider()
                    st.header("🔗 Reference Operations")
                    st.info("Use data from other sheets to enhance your current data")

                    ref_col1, ref_col2, ref_col3 = st.columns(3)

                    with ref_col1:
                        with st.expander("🔍 VLOOKUP from Reference"):
                            ref_source = st.selectbox(
                                "Reference source",
                                [s for s in st.session_state.data_sources.keys() if s != source_name],
                                key="vlookup_source"
                            )
                            if ref_source:
                                ref_df = st.session_state.data_sources[ref_source]
                                col1, col2 = st.columns(2)
                                with col1:
                                    lookup_col = st.selectbox("Lookup column in current data", df.columns, key="vlookup_col")
                                with col2:
                                    lookup_on = st.selectbox("Lookup column in reference", ref_df.columns, key="vlookup_on")
                                return_col = st.selectbox("Column to return from reference", ref_df.columns, key="vlookup_return")
                                new_col_name = st.text_input("New column name", f"{lookup_col}_from_{ref_source}", key="vlookup_newcol")

                                if st.button("✅ Apply VLOOKUP", use_container_width=True, key="apply_vlookup_btn"):
                                    result = MultiSheetProcessor.vlookup(df, lookup_col, ref_df, lookup_on, return_col)
                                    if result is not None:
                                        push_undo_snapshot()
                                        st.session_state.operations.append({
                                            'type': 'vlookup', 'lookup_source': ref_source,
                                            'lookup_col': lookup_col, 'lookup_on': lookup_on,
                                            'return_col': return_col, 'new_column': new_col_name
                                        })
                                        recompute_active_source()
                                        st.success(f"✅ Added column: {new_col_name}")
                                        st.rerun()

                    with ref_col2:
                        with st.expander("🎯 Filter by Reference"):
                            ref_source2 = st.selectbox(
                                "Reference source",
                                [s for s in st.session_state.data_sources.keys() if s != source_name],
                                key="filter_ref_source"
                            )
                            if ref_source2:
                                ref_df2 = st.session_state.data_sources[ref_source2]
                                col1, col2 = st.columns(2)
                                with col1:
                                    filter_col = st.selectbox("Column to filter", df.columns, key="filter_col")
                                with col2:
                                    ref_col = st.selectbox("Reference column", ref_df2.columns, key="filter_ref_col")
                                filter_op = st.selectbox("Operation", ['in', 'not in'], key="filter_ref_op")

                                if st.button("✅ Apply Filter", use_container_width=True, key="apply_filter_ref_btn"):
                                    result = MultiSheetProcessor.filter_by_reference(df, filter_col, ref_df2, ref_col, filter_op)
                                    if len(result) < len(df):
                                        push_undo_snapshot()
                                        st.session_state.operations.append({
                                            'type': 'filter_by_ref', 'ref_source': ref_source2,
                                            'column': filter_col, 'ref_column': ref_col, 'operation': filter_op
                                        })
                                        recompute_active_source()
                                        st.success(f"✅ Filtered from {len(df)} to {len(result)} rows")
                                        st.rerun()
                                    else:
                                        st.info("No rows filtered")

                    with ref_col3:
                        with st.expander("📋 Copy Column From Reference"):
                            st.caption(
                                "Pick a column from another loaded sheet and place it directly onto the "
                                "current data — no matching key needed, rows line up by position."
                            )
                            ref_source3 = st.selectbox(
                                "Reference source",
                                [s for s in st.session_state.data_sources.keys() if s != source_name],
                                key="copycol_source"
                            )
                            if ref_source3:
                                ref_df3 = st.session_state.data_sources[ref_source3]
                                copy_col = st.selectbox("Column to copy", ref_df3.columns, key="copycol_col")
                                new_copy_name = st.text_input(
                                    "New column name", copy_col, key="copycol_newname"
                                )
                                if len(ref_df3) != len(df):
                                    st.warning(
                                        f"⚠️ '{source_name}' has {len(df)} rows but '{ref_source3}' has "
                                        f"{len(ref_df3)}. Rows beyond the shorter sheet will be left blank — "
                                        "if these should line up by a shared ID instead, use VLOOKUP."
                                    )

                                if st.button("✅ Place Column", use_container_width=True, key="apply_copycol_btn"):
                                    push_undo_snapshot()
                                    st.session_state.operations.append({
                                        'type': 'copy_column', 'ref_source': ref_source3,
                                        'ref_column': copy_col, 'new_column': new_copy_name
                                    })
                                    recompute_active_source()
                                    st.success(f"✅ Placed column: {new_copy_name}")
                                    st.rerun()

                # Template Output Builder
                st.divider()
                with st.expander("🧩 Build Output From Template", expanded=False):
                    st.markdown(
                        "Define the exact headers your output report needs, then tell each "
                        "column where its data comes from — including looking values up from "
                        "a reference file, just like `XLOOKUP` in Excel."
                    )

                    header_source = st.radio(
                        "Where do the template headers come from?",
                        ["Type them in", "Use headers from another loaded source", "Upload a template file"],
                        horizontal=True, key="tpl_header_source"
                    )

                    if header_source == "Type them in":
                        headers_text = st.text_input(
                            "Comma-separated headers", ", ".join(st.session_state.template_headers),
                            key="tpl_headers_text"
                        )
                        if st.button("Set Headers", key="tpl_set_headers_btn"):
                            st.session_state.template_headers = [h.strip() for h in headers_text.split(",") if h.strip()]
                            st.session_state.template_source_file = None
                            st.rerun()

                    elif header_source == "Use headers from another loaded source":
                        other_sources = [s for s in st.session_state.data_sources.keys() if s != source_name]
                        if other_sources:
                            pick_source = st.selectbox("Copy headers from", other_sources, key="tpl_copy_headers_source")
                            if st.button("Set Headers", key="tpl_set_headers_from_source_btn"):
                                st.session_state.template_headers = list(st.session_state.data_sources[pick_source].columns)
                                st.session_state.template_source_file = None
                                st.rerun()
                        else:
                            st.info("Load another source first to copy its headers.")

                    else:
                        tpl_file = st.file_uploader(
                            "Upload a file — only its header row is used", type=['csv', 'xlsx', 'xls'],
                            key="tpl_file_uploader"
                        )
                        if tpl_file is not None:
                            try:
                                tpl_ext = Path(tpl_file.name).suffix.lower()
                                tpl_bytes = tpl_file.getvalue()
                                tpl_sheet = None
                                if tpl_ext != '.csv':
                                    tpl_xl = pd.ExcelFile(io.BytesIO(tpl_bytes))
                                    tpl_sheet = tpl_xl.sheet_names[0]
                                    if len(tpl_xl.sheet_names) > 1:
                                        st.caption(f"Using the first sheet: '{tpl_sheet}'")

                                tpl_raw_preview = read_raw_preview_bytes(tpl_bytes, tpl_ext, tpl_sheet)
                                tpl_ambiguous, tpl_suggested_row = analyze_header_ambiguity(tpl_raw_preview)

                                if tpl_ambiguous:
                                    st.warning(
                                        "⚠️ This file's header doesn't look like it's in the first row "
                                        "(blank/title row(s) detected above the data)."
                                    )
                                    st.caption("Raw preview — row numbers start at 0:")
                                    tpl_preview_display = tpl_raw_preview.copy()
                                    tpl_preview_display.index = [f"Row {i}" for i in range(len(tpl_preview_display))]
                                    st.dataframe(tpl_preview_display, use_container_width=True)
                                    max_row = max(len(tpl_raw_preview) - 1, 0)
                                    tpl_header_row = st.number_input(
                                        "Which row number is the header?",
                                        min_value=0, max_value=max_row,
                                        value=min(tpl_suggested_row, max_row),
                                        key="tpl_header_row_choice",
                                        help="Row 0 is the very first row of the file."
                                    )
                                    st.caption(f"💡 Suggested: Row {tpl_suggested_row}")
                                else:
                                    tpl_header_row = 0

                                if tpl_ext == '.csv':
                                    header_cols = list(pd.read_csv(io.BytesIO(tpl_bytes), header=tpl_header_row, nrows=0).columns)
                                else:
                                    header_cols = list(pd.read_excel(
                                        io.BytesIO(tpl_bytes), sheet_name=tpl_sheet, header=tpl_header_row, nrows=0
                                    ).columns)

                                if st.button("Set Headers", key="tpl_set_headers_from_file_btn"):
                                    st.session_state.template_headers = header_cols
                                    if tpl_ext in ('.xlsx', '.xlsm'):
                                        st.session_state.template_source_file = {
                                            'bytes': tpl_bytes, 'sheet': tpl_sheet, 'header_row': tpl_header_row,
                                            'name': tpl_file.name
                                        }
                                    else:
                                        st.session_state.template_source_file = None
                                        st.info(
                                            "Note: styled output (matching this template's fonts/colors/table) "
                                            "needs an .xlsx template - this file is .xls/.csv, so headers were "
                                            "used but the output will use standard formatting."
                                        )
                                    st.rerun()
                            except Exception as e:
                                st.error(f"Couldn't read headers from that file: {e}")

                    if st.session_state.template_headers:
                        st.success(f"Template has {len(st.session_state.template_headers)} column(s): "
                                   f"{', '.join(st.session_state.template_headers)}")
                        st.caption(
                            "The output will have exactly these columns, in this order — nothing extra "
                            "gets added, and any template header with no rule below is left blank."
                        )
                        st.divider()

                        other_sources_for_map = [s for s in st.session_state.data_sources.keys() if s != source_name]
                        am_targets = [f"{source_name} (current data)"] + other_sources_for_map
                        am_choice = st.selectbox(
                            "Match template headers against", am_targets, key="tpl_automap_source_choice",
                            help="Pick which loaded file/sheet to compare template headers against. Matches "
                                 "against the current data become a Direct copy; matches against a different "
                                 "loaded source become a Lookup (a shared column is auto-detected as the join "
                                 "key where possible)."
                        )
                        am_col1, am_col2 = st.columns([2, 3])
                        with am_col1:
                            if st.button("🪄 Auto-map columns by name", use_container_width=True, key="tpl_automap_btn",
                                         help="Matches template headers to source columns with the same "
                                              "(or very similar) name."):
                                import difflib
                                matched, unmatched, needs_key = [], [], []
                                target_is_current = (am_choice == am_targets[0])
                                target_df = df if target_is_current else st.session_state.data_sources[am_choice]

                                target_lookup = {c.lower().strip(): c for c in target_df.columns}
                                for header in st.session_state.template_headers:
                                    h_norm = header.lower().strip()
                                    match = target_lookup.get(h_norm)
                                    if not match:
                                        close = difflib.get_close_matches(h_norm, list(target_lookup.keys()), n=1, cutoff=0.8)
                                        if close:
                                            match = target_lookup[close[0]]
                                    if not match:
                                        unmatched.append(header)
                                        continue

                                    if target_is_current:
                                        st.session_state[f"tpl_mode_{header}"] = "Direct copy from current data"
                                        st.session_state[f"tpl_direct_{header}"] = match
                                        matched.append(f"{header} ← {match}")
                                    else:
                                        df_cols_lower = {c.lower().strip(): c for c in df.columns}
                                        other_cols_lower = {
                                            c.lower().strip(): c for c in target_df.columns if c != match
                                        }
                                        shared = [k for k in df_cols_lower if k in other_cols_lower]
                                        id_like = [k for k in shared if 'id' in k]
                                        key_norm = (id_like or shared or [None])[0]

                                        st.session_state[f"tpl_mode_{header}"] = "Lookup from reference file (XLOOKUP)"
                                        st.session_state[f"tpl_ref_{header}"] = am_choice
                                        st.session_state[f"tpl_return_{header}"] = match
                                        if key_norm:
                                            st.session_state[f"tpl_key_{header}"] = df_cols_lower[key_norm]
                                            st.session_state[f"tpl_refkey_{header}"] = other_cols_lower[key_norm]
                                            matched.append(
                                                f"{header} ← {am_choice}.{match} (joined on {df_cols_lower[key_norm]})"
                                            )
                                        else:
                                            needs_key.append(f"{header} ← {am_choice}.{match}")

                                st.session_state["_tpl_automap_summary"] = {
                                    "matched": matched, "unmatched": unmatched, "needs_key": needs_key
                                }
                                st.rerun()
                        with am_col2:
                            summary = st.session_state.pop("_tpl_automap_summary", None)
                            if summary:
                                if summary["matched"]:
                                    st.success("Mapped: " + "; ".join(summary["matched"]))
                                if summary.get("needs_key"):
                                    st.warning(
                                        "Matched but couldn't find a shared column to join on — set the key "
                                        "manually in each: " + "; ".join(summary["needs_key"])
                                    )
                                if summary["unmatched"]:
                                    st.info("No close match, set manually: " + ", ".join(summary["unmatched"]))

                        st.markdown("**Configure each output column:**")

                        other_sources = [s for s in st.session_state.data_sources.keys() if s != source_name]
                        mode_options = ["Direct copy from current data", "Lookup from reference file (XLOOKUP)", "Formula", "Static value"]
                        format_options = ["No formatting (as-is)", "Date", "Number", "Text case/trim"]

                        for header in st.session_state.template_headers:
                            with st.expander(f"🔧 {header}"):
                                mode = st.selectbox(
                                    "How should this column be filled?", mode_options,
                                    key=f"tpl_mode_{header}"
                                )

                                if mode == mode_options[0]:
                                    col_options = list(df.columns)
                                    st.selectbox("Copy from column", col_options, key=f"tpl_direct_{header}")

                                elif mode == mode_options[1]:
                                    if not other_sources:
                                        st.warning("Load a reference file (another data source) to use a lookup here.")
                                    else:
                                        ref_source = st.selectbox("Reference source", other_sources, key=f"tpl_ref_{header}")
                                        ref_df = st.session_state.data_sources[ref_source]
                                        col1, col2 = st.columns(2)
                                        with col1:
                                            st.selectbox("Match using column (current data)", df.columns, key=f"tpl_key_{header}")
                                        with col2:
                                            st.selectbox("Match against column (reference)", ref_df.columns, key=f"tpl_refkey_{header}")
                                        st.selectbox("Return this column from reference", ref_df.columns, key=f"tpl_return_{header}")
                                        st.text_input("Default value if not found (optional)", key=f"tpl_default_{header}")

                                elif mode == mode_options[2]:
                                    st.text_area(
                                        "Formula (supports [Columns], XLOOKUP('Source','Key','Return'), IF/THEN/ELSE, etc.)",
                                        "[Column1]", key=f"tpl_formula_{header}"
                                    )

                                else:
                                    st.text_input("Static value for every row", key=f"tpl_static_{header}")

                                st.markdown("—")
                                st.caption(
                                    "**Output format** — controls how the final values look, regardless of how "
                                    "they were sourced above. Use this when the template needs a specific date "
                                    "layout, decimal places, currency symbol, or text case that differs from the "
                                    "source file's own formatting."
                                )
                                fmt_type = st.selectbox("Format this column as", format_options, key=f"tpl_fmt_type_{header}")

                                if fmt_type == "Date":
                                    pattern_choice = st.selectbox(
                                        "Date format",
                                        ["YYYY-MM-DD", "MM/DD/YYYY", "DD/MM/YYYY", "Month D, YYYY", "Custom"],
                                        key=f"tpl_fmt_datepattern_{header}"
                                    )
                                    pattern_map = {
                                        "YYYY-MM-DD": "%Y-%m-%d",
                                        "MM/DD/YYYY": "%m/%d/%Y",
                                        "DD/MM/YYYY": "%d/%m/%Y",
                                        "Month D, YYYY": "%B %-d, %Y",
                                    }
                                    if pattern_choice == "Custom":
                                        st.text_input(
                                            "Custom strftime pattern (e.g. %b %d, %Y)", "%Y-%m-%d",
                                            key=f"tpl_fmt_datecustom_{header}"
                                        )
                                    st.caption("Values that aren't recognizable dates are left unchanged.")

                                elif fmt_type == "Number":
                                    nc1, nc2 = st.columns(2)
                                    with nc1:
                                        st.number_input("Decimal places", min_value=0, max_value=10, value=2,
                                                         key=f"tpl_fmt_decimals_{header}")
                                    with nc2:
                                        st.checkbox("Use thousands separator (1,000)", value=False,
                                                    key=f"tpl_fmt_thousands_{header}")
                                    nc3, nc4 = st.columns(2)
                                    with nc3:
                                        st.text_input("Prefix (e.g. $)", "", key=f"tpl_fmt_prefix_{header}")
                                    with nc4:
                                        st.text_input("Suffix (e.g. %)", "", key=f"tpl_fmt_suffix_{header}")
                                    st.caption("Values that aren't numbers are left unchanged.")

                                elif fmt_type == "Text case/trim":
                                    st.selectbox("Case", ["As-is", "UPPERCASE", "lowercase", "Title Case"],
                                                 key=f"tpl_fmt_case_{header}")
                                    st.checkbox("Trim leading/trailing whitespace", value=True, key=f"tpl_fmt_trim_{header}")

                        if st.button("🏗️ Build Output From Template", use_container_width=True, key="tpl_build_btn"):
                            all_sources = st.session_state.data_sources
                            column_configs = {}
                            for header in st.session_state.template_headers:
                                mode = st.session_state.get(f"tpl_mode_{header}", mode_options[0])
                                if mode == mode_options[0]:
                                    column_configs[header] = {
                                        'mode': 'direct',
                                        'source_column': st.session_state.get(f"tpl_direct_{header}")
                                    }
                                elif mode == mode_options[1]:
                                    column_configs[header] = {
                                        'mode': 'lookup',
                                        'ref_source': st.session_state.get(f"tpl_ref_{header}"),
                                        'key_column': st.session_state.get(f"tpl_key_{header}"),
                                        'ref_key_column': st.session_state.get(f"tpl_refkey_{header}"),
                                        'ref_return_column': st.session_state.get(f"tpl_return_{header}"),
                                        'default': st.session_state.get(f"tpl_default_{header}") or None
                                    }
                                elif mode == mode_options[2]:
                                    formula_text = st.session_state.get(f"tpl_formula_{header}", "")
                                    column_configs[header] = {
                                        'mode': 'formula',
                                        'formula': formula_text,
                                        'referenced_sources': detect_referenced_source_names(formula_text, all_sources.keys())
                                    }
                                else:
                                    column_configs[header] = {
                                        'mode': 'static',
                                        'value': st.session_state.get(f"tpl_static_{header}", '')
                                    }

                                fmt_type = st.session_state.get(f"tpl_fmt_type_{header}", format_options[0])
                                if fmt_type == "Date":
                                    pattern_choice = st.session_state.get(f"tpl_fmt_datepattern_{header}", "YYYY-MM-DD")
                                    pattern_map = {
                                        "YYYY-MM-DD": "%Y-%m-%d",
                                        "MM/DD/YYYY": "%m/%d/%Y",
                                        "DD/MM/YYYY": "%d/%m/%Y",
                                        "Month D, YYYY": "%B %-d, %Y",
                                    }
                                    pattern = st.session_state.get(f"tpl_fmt_datecustom_{header}", "%Y-%m-%d") \
                                        if pattern_choice == "Custom" else pattern_map.get(pattern_choice, "%Y-%m-%d")
                                    column_configs[header]['format'] = {'type': 'date', 'pattern': pattern}
                                elif fmt_type == "Number":
                                    column_configs[header]['format'] = {
                                        'type': 'number',
                                        'decimals': st.session_state.get(f"tpl_fmt_decimals_{header}", 2),
                                        'thousands': st.session_state.get(f"tpl_fmt_thousands_{header}", False),
                                        'prefix': st.session_state.get(f"tpl_fmt_prefix_{header}", ""),
                                        'suffix': st.session_state.get(f"tpl_fmt_suffix_{header}", ""),
                                    }
                                elif fmt_type == "Text case/trim":
                                    case_map = {"As-is": "as_is", "UPPERCASE": "upper", "lowercase": "lower", "Title Case": "title"}
                                    column_configs[header]['format'] = {
                                        'type': 'text',
                                        'case': case_map.get(st.session_state.get(f"tpl_fmt_case_{header}", "As-is"), "as_is"),
                                        'trim': st.session_state.get(f"tpl_fmt_trim_{header}", True),
                                    }
                                else:
                                    column_configs[header]['format'] = {'type': 'none'}

                            result_df = MultiSheetProcessor.build_template_output(
                                df, st.session_state.template_headers, column_configs, all_sources
                            )
                            result_name = f"{source_name}_template_output"
                            st.session_state.data_sources[result_name] = result_df
                            st.session_state.source_metadata[result_name] = {
                                'file': 'Template Output', 'sheet': 'Built',
                                'rows': len(result_df), 'columns': len(result_df.columns),
                                'template_source_file': st.session_state.template_source_file
                            }
                            build_op = {
                                'type': 'build_template',
                                'template_headers': list(st.session_state.template_headers),
                                'column_configs': column_configs
                            }
                            st.session_state.source_snapshots[result_name] = df.copy()
                            st.session_state.operations = [build_op]
                            st.session_state.operations_undo_stack = []
                            st.session_state.active_source = result_name
                            if st.session_state.template_source_file:
                                st.success(
                                    f"✅ Built output '{result_name}' with {len(result_df)} rows. "
                                    "Head to the **📥 Export** operation to download it styled to match "
                                    "your uploaded template's font/color/table formatting."
                                )
                            else:
                                st.success(f"✅ Built output '{result_name}' with {len(result_df)} rows")
                            st.rerun()

                # Standard Operations Toolbar
                st.divider()
                st.header("⚡ Data Operations")

                if st.session_state.current_operation:
                    col1, col2 = st.columns([6, 1])
                    with col2:
                        if st.button("❌ Close", use_container_width=True, key="close_op_btn"):
                            st.session_state.current_operation = None
                            st.rerun()

                if not st.session_state.current_operation:
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        if st.button("🔍 Filter Rows", use_container_width=True, key="op_filter_btn"):
                            st.session_state.current_operation = "filter"; st.rerun()
                        if st.button("🔍 Multiple Filters", use_container_width=True, key="op_multi_filter_btn"):
                            st.session_state.current_operation = "filter_multiple"; st.rerun()
                        if st.button("📐 Add Column", use_container_width=True, key="op_calc_btn"):
                            st.session_state.current_operation = "calc"; st.rerun()
                        if st.button("📊 Group By", use_container_width=True, key="op_group_btn"):
                            st.session_state.current_operation = "group"; st.rerun()
                        if st.button("🧹 Remove Duplicates", use_container_width=True, key="op_dedupe_btn"):
                            st.session_state.current_operation = "dedupe"; st.rerun()
                    with col2:
                        if st.button("🔃 Sort Data", use_container_width=True, key="op_sort_btn"):
                            st.session_state.current_operation = "sort"; st.rerun()
                        if st.button("✏️ Rename Column", use_container_width=True, key="op_rename_btn"):
                            st.session_state.current_operation = "rename"; st.rerun()
                        if st.button("🗑️ Drop Column", use_container_width=True, key="op_drop_btn"):
                            st.session_state.current_operation = "drop"; st.rerun()
                        if st.button("📑 Select Columns to Keep", use_container_width=True, key="op_select_cols_btn"):
                            st.session_state.current_operation = "select_columns"; st.rerun()
                        if st.button("✂️ Split Column", use_container_width=True, key="op_split_btn"):
                            st.session_state.current_operation = "split_column"; st.rerun()
                    with col3:
                        if st.button("📥 Append Data", use_container_width=True, key="op_append_btn"):
                            st.session_state.current_operation = "append"; st.rerun()
                        if st.button("📈 Pivot Table", use_container_width=True, key="op_pivot_btn"):
                            st.session_state.current_operation = "pivot"; st.rerun()
                        if st.button("📊 Summarize by Ref", use_container_width=True, key="op_summarize_btn"):
                            st.session_state.current_operation = "summarize"; st.rerun()
                        if st.button("🔎 Find & Replace", use_container_width=True, key="op_find_replace_btn"):
                            st.session_state.current_operation = "find_replace"; st.rerun()
                    with col4:
                        if st.button("💾 Execute All", use_container_width=True, key="op_execute_btn"):
                            st.session_state.current_operation = "execute"; st.rerun()
                        if st.button("📥 Export", use_container_width=True, key="op_export_btn"):
                            st.session_state.current_operation = "export"; st.rerun()
                        if st.button("🩹 Handle Missing Values", use_container_width=True, key="op_missing_btn"):
                            st.session_state.current_operation = "missing"; st.rerun()
                        if st.button("🔤 Change Data Type", use_container_width=True, key="op_dtype_btn"):
                            st.session_state.current_operation = "dtype"; st.rerun()

                if st.session_state.current_operation:
                    st.divider()
                    st.markdown('<div class="operation-active">', unsafe_allow_html=True)
                    st.markdown(f"### 🔧 Operation: {st.session_state.current_operation.title()}")

                    op_selected = st.session_state.current_operation

                    if op_selected == "filter":
                        col1, col2 = st.columns(2)
                        with col1:
                            column = st.selectbox("Select column to filter", df.columns, key="filter_column")
                            dtype = df[column].dtype
                            st.caption(f"Data type: {dtype}")
                            unique_count = df[column].nunique()
                            st.caption(f"Unique values: {unique_count}")
                            
                        with col2:
                            operation = st.selectbox("Operation", ['==', '!=', '>', '<', '>=', '<=', 'contains'], key="filter_operation")
                            
                            # Check if this is a date column
                            is_date_column = pd.api.types.is_datetime64_any_dtype(dtype) or try_convert_datetime_column(df[column]) is not None
                            
                            if is_date_column:
                                # Date column with shortcuts
                                st.markdown('<div class="date-shortcut-box">', unsafe_allow_html=True)
                                st.caption("📅 Date Selection Options:")
                                
                                date_mode = st.radio(
                                    "Select date mode:",
                                    ["Exact Date", "Today", "Yesterday", "This Week", "This Month", "Last Month", "Last 7 Days", "Last 30 Days", "Last 90 Days", "Year to Date"],
                                    horizontal=False,
                                    key="filter_date_mode",
                                    index=0
                                )
                                
                                if date_mode == "Exact Date":
                                    min_date = df[column].min()
                                    max_date = df[column].max()
                                    st.caption(f"Date range: {min_date.date()} to {max_date.date()}")
                                    min_d = min_date.date() if pd.notna(min_date) else None
                                    max_d = max_date.date() if pd.notna(max_date) else None
                                    today = pd.Timestamp.now().date()
                                    if min_d is not None and today < min_d:
                                        default_d = min_d
                                    elif max_d is not None and today > max_d:
                                        default_d = max_d
                                    else:
                                        default_d = today
                                    value = st.date_input(
                                        "Select date", 
                                        value=default_d,
                                        min_value=min_d,
                                        max_value=max_d,
                                        key="filter_date_value"
                                    )
                                    value = value.strftime("%Y-%m-%d")
                                    shortcut = 'exact'
                                else:
                                    shortcut_map = {
                                        "Today": "today",
                                        "Yesterday": "yesterday",
                                        "This Week": "this_week",
                                        "This Month": "this_month",
                                        "Last Month": "last_month",
                                        "Last 7 Days": "last_7_days",
                                        "Last 30 Days": "last_30_days",
                                        "Last 90 Days": "last_90_days",
                                        "Year to Date": "year_to_date"
                                    }
                                    shortcut = shortcut_map.get(date_mode, 'exact')
                                    start_date, end_date = get_date_from_shortcut(shortcut)
                                    if start_date and end_date:
                                        st.info(f"📆 Filtering for: **{start_date.strftime('%Y-%m-%d')}** to **{end_date.strftime('%Y-%m-%d')}**")
                                    value = None
                                st.markdown('</div>', unsafe_allow_html=True)
                                
                            elif dtype == 'object' or pd.api.types.is_string_dtype(dtype):
                                # Text column
                                unique_values, has_more = get_unique_values_sorted(df, column)
                                if unique_values:
                                    if has_more:
                                        st.caption(f"Showing first {len(unique_values)} of {df[column].nunique()} values")
                                    value = st.selectbox("Select value to filter", options=unique_values, key="filter_value_dropdown")
                                else:
                                    value = st.text_input("Enter value to filter", key="filter_value_text")
                                shortcut = 'exact'
                            
                            elif pd.api.types.is_numeric_dtype(dtype):
                                # Numeric column
                                min_val = df[column].min()
                                max_val = df[column].max()
                                st.caption(f"Range: {min_val} to {max_val}")
                                
                                if operation in ['==', '!=']:
                                    unique_values, has_more = get_unique_values_sorted(df, column)
                                    if len(unique_values) <= 50:
                                        if has_more:
                                            st.caption(f"Showing first {len(unique_values)} of {df[column].nunique()} values")
                                        value = st.selectbox("Select value", options=unique_values, key="filter_numeric_dropdown")
                                    else:
                                        value = st.number_input("Enter value", value=float(min_val) if pd.notna(min_val) else 0.0, key="filter_numeric_input")
                                else:
                                    value = st.number_input("Enter value", value=float(min_val) if pd.notna(min_val) else 0.0, key="filter_numeric_input")
                                shortcut = 'exact'
                            else:
                                value = st.text_input("Enter value to filter", key="filter_value_fallback")
                                shortcut = 'exact'

                        if st.button("✅ Apply Filter", use_container_width=True, key="apply_filter_btn"):
                            # Apply the filter with date shortcut support
                            if shortcut != 'exact' and is_date_column:
                                result_df = apply_date_shortcut_to_filter(df, column, operation, shortcut)
                            else:
                                if operation == 'contains':
                                    value = str(value)
                                result_df = MultiSheetProcessor.filter_rows(df, column, operation, value)
                            
                            if len(result_df) < len(df):
                                push_undo_snapshot()
                                op_data = {
                                    'type': 'filter', 
                                    'column': column, 
                                    'operation': operation,
                                    'date_shortcut': shortcut if shortcut != 'exact' else None
                                }
                                if shortcut == 'exact':
                                    op_data['value'] = str(value)
                                else:
                                    start_d, end_d = get_date_from_shortcut(shortcut)
                                    op_data['value'] = f"{start_d.strftime('%Y-%m-%d')} to {end_d.strftime('%Y-%m-%d')}"
                                    op_data['shortcut_label'] = get_date_filter_description(shortcut)
                                st.session_state.operations.append(op_data)
                                recompute_active_source()
                                st.success(f"✅ Filtered from {len(df)} to {len(result_df)} rows")
                                st.rerun()
                            else:
                                st.warning("No rows were filtered. Check your filter criteria.")

                    elif op_selected == "filter_multiple":
                        st.info(
                            "Add multiple filter conditions, then choose how they combine. "
                            "**AND** = a row must match every condition. "
                            "**OR** = a row must match at least one condition."
                        )
                        st.session_state.filter_logic = st.radio(
                            "Combine conditions using",
                            ["AND", "OR"],
                            index=0 if st.session_state.filter_logic == "AND" else 1,
                            horizontal=True,
                            key="multi_filter_logic_radio"
                        )

                        col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                        with col1:
                            new_column = st.selectbox("Column", df.columns, key="multi_filter_col")
                        with col2:
                            ops_list = ['==', '!=', '>', '<', '>=', '<=', 'contains']
                            new_operation = st.selectbox("Operation", ops_list, key="multi_filter_op")
                        with col3:
                            dtype = df[new_column].dtype
                            is_date_column = pd.api.types.is_datetime64_any_dtype(dtype) or try_convert_datetime_column(df[new_column]) is not None
                            
                            if is_date_column:
                                date_mode = st.selectbox(
                                    "Date mode",
                                    ["Exact Date", "Today", "Yesterday", "This Week", "This Month", "Last Month", "Last 7 Days", "Last 30 Days"],
                                    key=f"multi_filter_date_mode_{len(st.session_state.filter_conditions)}"
                                )
                                if date_mode == "Exact Date":
                                    min_date = df[new_column].min()
                                    max_date = df[new_column].max()
                                    min_d = min_date.date() if pd.notna(min_date) else None
                                    max_d = max_date.date() if pd.notna(max_date) else None
                                    today = pd.Timestamp.now().date()
                                    if min_d is not None and today < min_d:
                                        default_d = min_d
                                    elif max_d is not None and today > max_d:
                                        default_d = max_d
                                    else:
                                        default_d = today
                                    new_value = st.date_input(
                                        "Date", 
                                        value=default_d,
                                        min_value=min_d,
                                        max_value=max_d,
                                        key="multi_filter_date"
                                    )
                                    new_value = new_value.strftime("%Y-%m-%d")
                                else:
                                    shortcut_map = {
                                        "Today": "today",
                                        "Yesterday": "yesterday",
                                        "This Week": "this_week",
                                        "This Month": "this_month",
                                        "Last Month": "last_month",
                                        "Last 7 Days": "last_7_days",
                                        "Last 30 Days": "last_30_days"
                                    }
                                    shortcut = shortcut_map.get(date_mode)
                                    start_date, end_date = get_date_from_shortcut(shortcut)
                                    new_value = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}" if start_date else ""
                                    st.info(f"📆 {date_mode}: {new_value}")
                            elif dtype == 'object' or pd.api.types.is_string_dtype(dtype):
                                unique_values, has_more = get_unique_values_sorted(df, new_column)
                                if unique_values:
                                    new_value = st.selectbox("Value", options=unique_values, key="multi_filter_text_dropdown")
                                else:
                                    new_value = st.text_input("Value", key="multi_filter_text")
                            elif pd.api.types.is_numeric_dtype(dtype):
                                min_val = df[new_column].min()
                                max_val = df[new_column].max()
                                unique_values, has_more = get_unique_values_sorted(df, new_column)
                                if len(unique_values) <= 50:
                                    new_value = st.selectbox("Value", options=unique_values, key="multi_filter_num_dropdown")
                                else:
                                    new_value = st.number_input("Value", value=float(min_val) if pd.notna(min_val) else 0.0, key="multi_filter_num")
                            else:
                                new_value = st.text_input("Value", key="multi_filter_fallback")
                        with col4:
                            if st.button("➕ Add", key="add_filter_condition"):
                                if new_column and new_value is not None:
                                    st.session_state.filter_conditions.append({
                                        'column': new_column,
                                        'operation': new_operation,
                                        'value': str(new_value)
                                    })
                                    st.rerun()
                        
                        if st.session_state.filter_conditions:
                            joiner = f" {st.session_state.filter_logic} "
                            st.write(f"Current filter conditions (combined with **{st.session_state.filter_logic}**):")
                            for i, cond in enumerate(st.session_state.filter_conditions):
                                col1, col2 = st.columns([5, 1])
                                with col1:
                                    prefix = "" if i == 0 else f"{st.session_state.filter_logic} "
                                    st.write(f"{i+1}. {prefix}{cond['column']} {cond['operation']} {cond['value']}")
                                with col2:
                                    if st.button("🗑️", key=f"remove_filter_{i}"):
                                        st.session_state.filter_conditions.pop(i)
                                        st.rerun()
                            
                            if st.button("✅ Apply All Filters", use_container_width=True):
                                logic = st.session_state.filter_logic
                                new_df = MultiSheetProcessor.filter_multiple_rows(df, st.session_state.filter_conditions, logic)
                                push_undo_snapshot()
                                st.session_state.operations.append({
                                    'type': 'filter_multiple',
                                    'conditions': st.session_state.filter_conditions.copy(),
                                    'logic': logic
                                })
                                st.session_state.filter_conditions = []
                                recompute_active_source()
                                st.success(f"✅ Filtered from {len(df)} to {len(new_df)} rows")
                                st.rerun()
                        else:
                            st.info("Add at least one filter condition above")

                    elif op_selected == "calc":
                        st.info("""
                        **Formula Examples:**
                        - `[Sales] * 0.1` (10% commission)
                        - `([Price] - [Cost]) / [Cost] * 100` (profit margin %)
                        - `[First Name] + ' ' + [Last Name]` (full name)
                        - `IF [Sales] > 1000 THEN 'High' ELSE 'Low' END`
                        - `abs([Value])` (absolute value)
                        - `round([Price], 2)` (round to 2 decimals)
                        - `upper([Name])` / `lower([Name])`
                        - `XLOOKUP([Product ID], 'Products', 'ID', 'Unit Price')` — Excel-XLOOKUP-style
                          lookup into any other loaded sheet/file. Optional 5th argument is a
                          fallback value for no match, e.g. `XLOOKUP([SKU], 'Catalog', 'SKU', 'Price', 0)`
                        """)
                        new_column = st.text_input("New column name", f"calculated_{len(df.columns)}", key="calc_newcol")
                        formula = st.text_area("Formula", "[Column1] + [Column2]", key="calc_formula")

                        if st.button("✅ Add Column", use_container_width=True, key="apply_calc_btn"):
                            if new_column and formula:
                                all_sources = st.session_state.data_sources
                                new_df = MultiSheetProcessor.add_calculated_column(df, new_column, formula, all_sources)
                                if new_column in new_df.columns:
                                    push_undo_snapshot()
                                    referenced = detect_referenced_source_names(formula, all_sources.keys())
                                    st.session_state.operations.append({
                                        'type': 'calc_column', 'new_column': new_column, 'formula': formula,
                                        'referenced_sources': referenced
                                    })
                                    recompute_active_source()
                                    st.success(f"✅ Added column: {new_column}")
                                    st.rerun()
                            else:
                                st.warning("Please provide both column name and formula")

                    elif op_selected == "group":
                        if len(df.columns) < 2:
                            st.warning("Group By needs at least 2 columns (one to group by, one to aggregate).")
                            st.stop()
                        col1, col2 = st.columns(2)
                        with col1:
                            group_col = st.selectbox("Group by column", df.columns, key="group_by_col")
                        with col2:
                            agg_col = st.selectbox("Column to aggregate", [c for c in df.columns if c != group_col], key="group_agg_col")
                        agg_func = st.selectbox("Aggregation function", ['sum', 'mean', 'median', 'min', 'max', 'count', 'std'], key="group_agg_func")

                        if st.button("✅ Apply Group By", use_container_width=True, key="apply_group_btn"):
                            new_df = MultiSheetProcessor.group_by(df, group_col, agg_col, agg_func)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'group_by', 'group_column': group_col,
                                'aggregate_column': agg_col, 'function': agg_func
                            })
                            recompute_active_source()
                            st.success(f"✅ Grouped data: {len(new_df)} groups")
                            st.rerun()

                    elif op_selected == "sort":
                        sort_col = st.selectbox("Sort by column", df.columns, key="sort_col")
                        ascending = st.radio("Order", ["Ascending", "Descending"], horizontal=True, key="sort_order")

                        if st.button("✅ Apply Sort", use_container_width=True, key="apply_sort_btn"):
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'sort', 'column': sort_col, 'ascending': ascending == "Ascending"
                            })
                            recompute_active_source()
                            st.success(f"✅ Sorted by {sort_col}")
                            st.rerun()

                    elif op_selected == "rename":
                        old_name = st.selectbox("Select column to rename", df.columns, key="rename_old")
                        new_name = st.text_input("New name", f"{old_name}_new", key="rename_new")

                        if st.button("✅ Apply Rename", use_container_width=True, key="apply_rename_btn"):
                            if new_name and new_name not in df.columns:
                                push_undo_snapshot()
                                st.session_state.operations.append({
                                    'type': 'rename', 'old_name': old_name, 'new_name': new_name
                                })
                                recompute_active_source()
                                st.success(f"✅ Renamed '{old_name}' to '{new_name}'")
                                st.rerun()
                            else:
                                st.warning("New name is empty or already exists")

                    elif op_selected == "drop":
                        col_to_drop = st.selectbox("Select column to drop", df.columns, key="drop_col")
                        confirmed = st.checkbox(f"Confirm dropping '{col_to_drop}'", key="drop_confirm")

                        if st.button("✅ Apply Drop", use_container_width=True, key="apply_drop_btn", disabled=not confirmed):
                            push_undo_snapshot()
                            st.session_state.operations.append({'type': 'drop', 'column': col_to_drop})
                            recompute_active_source()
                            st.success(f"✅ Dropped column: {col_to_drop}")
                            st.rerun()

                    elif op_selected == "select_columns":
                        st.info(
                            "Pick the exact set of columns you want to **keep** — everything else gets "
                            "dropped in one step. This is the quickest way to extract just the fields "
                            "you need. The order you select columns in is the order they'll appear in "
                            "the output."
                        )
                        selected_cols = st.multiselect(
                            "Columns to keep", options=list(df.columns), default=list(df.columns),
                            key="select_columns_multiselect",
                            help="Deselect anything you don't want in the output."
                        )
                        c1, c2 = st.columns(2)
                        with c1:
                            if st.button("Select All", use_container_width=True, key="select_cols_all_btn"):
                                st.session_state["select_columns_multiselect"] = list(df.columns)
                                st.rerun()
                        with c2:
                            if st.button("Clear All", use_container_width=True, key="select_cols_clear_btn"):
                                st.session_state["select_columns_multiselect"] = []
                                st.rerun()

                        st.caption(f"Keeping {len(selected_cols)} of {len(df.columns)} column(s).")

                        if st.button("✅ Apply Column Selection", use_container_width=True, key="apply_select_cols_btn",
                                     disabled=(len(selected_cols) == 0)):
                            new_df = MultiSheetProcessor.select_columns(df, selected_cols)
                            push_undo_snapshot()
                            st.session_state.operations.append({'type': 'select_columns', 'columns': list(selected_cols)})
                            recompute_active_source()
                            st.success(f"✅ Kept {len(selected_cols)} column(s), dropped {len(df.columns) - len(selected_cols)}")
                            st.rerun()

                    elif op_selected == "dedupe":
                        st.info("Remove duplicate rows — optionally based on only some columns (e.g. dedupe by "
                                 "Email even if other columns differ).")
                        subset_cols = st.multiselect(
                            "Consider only these columns (leave empty = compare full rows)",
                            list(df.columns), key="dedupe_subset"
                        )
                        keep_choice = st.radio(
                            "Which copy to keep", ["First occurrence", "Last occurrence", "Drop all duplicates"],
                            horizontal=True, key="dedupe_keep"
                        )
                        keep_map = {"First occurrence": "first", "Last occurrence": "last", "Drop all duplicates": "drop_all"}

                        if st.button("✅ Remove Duplicates", use_container_width=True, key="apply_dedupe_btn"):
                            subset = subset_cols if subset_cols else None
                            keep_val = keep_map[keep_choice]
                            new_df = MultiSheetProcessor.remove_duplicates(df, subset, keep_val)
                            push_undo_snapshot()
                            st.session_state.operations.append({'type': 'remove_duplicates', 'subset': subset, 'keep': keep_val})
                            recompute_active_source()
                            st.success(f"✅ Removed {len(df) - len(new_df)} duplicate row(s). {len(new_df)} row(s) remain.")
                            st.rerun()

                    elif op_selected == "missing":
                        st.info("Fill in or drop rows with missing (blank/NaN) values.")
                        target_col = st.selectbox("Column", ["All columns"] + list(df.columns), key="missing_col")
                        method_choice = st.selectbox(
                            "Method",
                            ["Drop rows with missing values", "Fill with a specific value",
                             "Fill with column mean (numeric)", "Fill with column median (numeric)",
                             "Fill with most common value (mode)", "Fill forward (carry last value down)",
                             "Fill backward (carry next value up)"],
                            key="missing_method"
                        )
                        method_map = {
                            "Drop rows with missing values": "drop_rows", "Fill with a specific value": "fill_static",
                            "Fill with column mean (numeric)": "fill_mean", "Fill with column median (numeric)": "fill_median",
                            "Fill with most common value (mode)": "fill_mode",
                            "Fill forward (carry last value down)": "fill_forward",
                            "Fill backward (carry next value up)": "fill_backward",
                        }
                        fill_value = None
                        if method_choice == "Fill with a specific value":
                            fill_value = st.text_input("Value to fill with", key="missing_fill_value")

                        if st.button("✅ Apply", use_container_width=True, key="apply_missing_btn"):
                            col_val = None if target_col == "All columns" else target_col
                            method_val = method_map[method_choice]
                            new_df = MultiSheetProcessor.handle_missing_values(df, col_val, method_val, fill_value)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'handle_missing', 'column': col_val, 'method': method_val, 'value': fill_value
                            })
                            recompute_active_source()
                            st.success(f"✅ Done. {len(new_df)} row(s) remain.")
                            st.rerun()

                    elif op_selected == "dtype":
                        st.info("Convert a column to a different data type. Values that can't convert become "
                                 "blank rather than breaking the whole column.")
                        target_col = st.selectbox("Column", df.columns, key="dtype_col")
                        st.caption(f"Current type: {df[target_col].dtype}")
                        type_choice = st.selectbox(
                            "Convert to",
                            ["Text", "Whole Number (Integer)", "Decimal Number", "Date", "Yes/No (Boolean)"],
                            key="dtype_target"
                        )
                        type_map = {
                            "Text": "text", "Whole Number (Integer)": "integer", "Decimal Number": "float",
                            "Date": "date", "Yes/No (Boolean)": "boolean"
                        }

                        if st.button("✅ Convert", use_container_width=True, key="apply_dtype_btn"):
                            target_type = type_map[type_choice]
                            new_df = MultiSheetProcessor.change_data_type(df, target_col, target_type)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'change_dtype', 'column': target_col, 'target_type': target_type
                            })
                            recompute_active_source()
                            st.success(f"✅ Converted '{target_col}' to {type_choice}")
                            st.rerun()

                    elif op_selected == "find_replace":
                        st.info("Find and replace text within a column, or across every text column at once.")
                        target_col = st.selectbox("Column", ["All text columns"] + list(df.columns), key="fr_col")
                        find_text = st.text_input("Find", key="fr_find")
                        replace_text = st.text_input("Replace with", key="fr_replace")
                        c1, c2 = st.columns(2)
                        with c1:
                            match_case = st.checkbox("Match case", value=False, key="fr_case")
                        with c2:
                            whole_cell = st.checkbox("Whole cell must match exactly (not just contain)", value=False, key="fr_whole")

                        if st.button("✅ Replace", use_container_width=True, key="apply_fr_btn", disabled=not find_text):
                            col_val = None if target_col == "All text columns" else target_col
                            new_df = MultiSheetProcessor.find_replace(df, col_val, find_text, replace_text, match_case, whole_cell)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'find_replace', 'column': col_val, 'find_text': find_text,
                                'replace_text': replace_text, 'match_case': match_case, 'whole_cell': whole_cell
                            })
                            recompute_active_source()
                            st.success("✅ Find & Replace applied")
                            st.rerun()

                    elif op_selected == "split_column":
                        st.info("Split one column into multiple columns using a delimiter — e.g. split "
                                 "'Full Name' by a space into first/last name columns.")
                        target_col = st.selectbox("Column to split", df.columns, key="split_col")
                        delimiter = st.text_input("Delimiter", " ", key="split_delim")
                        new_names = None
                        if delimiter:
                            preview_split = df[target_col].astype(str).str.split(delimiter, expand=True).head(5)
                            st.caption("Preview:")
                            st.dataframe(preview_split, use_container_width=True)
                            n_parts = preview_split.shape[1]
                            names_text = st.text_input(
                                f"New column names ({n_parts}, comma-separated — leave blank for auto names)",
                                key="split_names"
                            )
                            new_names = [n.strip() for n in names_text.split(",") if n.strip()] if names_text else None

                        if st.button("✅ Split Column", use_container_width=True, key="apply_split_btn", disabled=not delimiter):
                            new_df = MultiSheetProcessor.split_column(df, target_col, delimiter, new_names)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'split_column', 'column': target_col, 'delimiter': delimiter, 'new_names': new_names
                            })
                            recompute_active_source()
                            st.success(f"✅ Split '{target_col}' into {new_df.shape[1] - df.shape[1] + 1} new column(s)")
                            st.rerun()

                    elif op_selected == "append":
                        st.info("Append rows from another data source to the current data")
                        append_source = st.selectbox(
                            "Source to append",
                            [s for s in st.session_state.data_sources.keys() if s != source_name],
                            key="append_source_select"
                        )
                        if append_source:
                            append_df = st.session_state.data_sources[append_source]
                            st.write(f"Will append {len(append_df)} rows from '{append_source}'")
                            missing_cols = [c for c in df.columns if c not in append_df.columns]
                            if missing_cols:
                                st.warning(f"Missing columns in append source: {missing_cols}")
                                st.write("They will be filled with NaN")

                            if st.button("✅ Append Data", use_container_width=True, key="apply_append_btn"):
                                new_df = MultiSheetProcessor.append_data(df, append_df)
                                push_undo_snapshot()
                                st.session_state.operations.append({'type': 'append', 'append_source': append_source})
                                recompute_active_source()
                                st.success(f"✅ Appended {len(append_df)} rows. New total: {len(new_df)}")
                                st.rerun()

                    elif op_selected == "pivot":
                        if len(df.columns) < 3:
                            st.warning("Pivot Table needs at least 3 columns (index, columns, and values).")
                            st.stop()
                        col1, col2 = st.columns(2)
                        with col1:
                            index = st.selectbox("Index column", df.columns, key="pivot_index")
                            columns = st.selectbox("Columns to pivot", [c for c in df.columns if c != index], key="pivot_columns")
                        with col2:
                            values = st.selectbox("Values column", [c for c in df.columns if c not in [index, columns]], key="pivot_values")
                            aggfunc = st.selectbox("Aggregation", ['sum', 'mean', 'count', 'min', 'max'], key="pivot_aggfunc")

                        if st.button("✅ Create Pivot Table", use_container_width=True, key="apply_pivot_btn"):
                            new_df = MultiSheetProcessor.pivot_table(df, index, columns, values, aggfunc)
                            push_undo_snapshot()
                            st.session_state.operations.append({
                                'type': 'pivot', 'index': index, 'columns': columns,
                                'values': values, 'aggfunc': aggfunc
                            })
                            recompute_active_source()
                            st.success(f"✅ Pivot table created: {len(new_df)} rows")
                            st.rerun()

                    elif op_selected == "summarize":
                        st.info("Summarize data and join with reference data")
                        ref_source = st.selectbox(
                            "Reference source",
                            [s for s in st.session_state.data_sources.keys() if s != source_name],
                            key="summarize_ref_source"
                        )
                        if ref_source:
                            ref_df = st.session_state.data_sources[ref_source]
                            col1, col2 = st.columns(2)
                            with col1:
                                group_col = st.selectbox("Group by column in current data", df.columns, key="summarize_group_col")
                                ref_col = st.selectbox("Reference column", ref_df.columns, key="summarize_ref_col")
                            with col2:
                                agg_col = st.selectbox("Column to aggregate", [c for c in df.columns if c != group_col], key="summarize_agg_col")
                                agg_func = st.selectbox("Aggregation function", ['sum', 'mean', 'count', 'min', 'max'], key="summarize_agg_func")

                            if st.button("✅ Summarize", use_container_width=True, key="apply_summarize_btn"):
                                new_df = MultiSheetProcessor.summarize_by_reference(df, group_col, agg_col, ref_df, ref_col, agg_func)
                                push_undo_snapshot()
                                st.session_state.operations.append({
                                    'type': 'summarize_by_ref', 'ref_source': ref_source,
                                    'group_col': group_col, 'agg_col': agg_col,
                                    'ref_col': ref_col, 'agg_func': agg_func
                                })
                                recompute_active_source()
                                st.success(f"✅ Summarized data: {len(new_df)} rows")
                                st.rerun()

                    elif op_selected == "execute":
                        if st.session_state.operations:
                            st.info(
                                "Each step already reruns automatically as you build. Use this if a "
                                "reference source changed and you want to refresh the result."
                            )
                            st.subheader(f"Current recipe on '{source_name}' ({len(st.session_state.operations)} step(s)):")
                            for i, op in enumerate(st.session_state.operations, 1):
                                st.write(f"{i}. {describe_operation(op)}")

                            if st.button("🔄 Recompute From Scratch", use_container_width=True, key="execute_all_btn"):
                                recompute_active_source()
                                new_df = st.session_state.data_sources[source_name]
                                st.success(f"✅ Recomputed! New shape: {new_df.shape}")
                                st.rerun()
                        else:
                            st.info("No operations to execute")

                    elif op_selected == "export":
                        st.subheader("📥 Export Data")
                        st.caption(
                            "This runs in your browser, so — like any web app — it can't reach into your "
                            "computer and overwrite your original file directly; that's a security boundary "
                            "every browser enforces, not a limitation specific to this tool. Exporting always "
                            "creates a new downloaded file."
                        )

                        original_file = meta.get('file')
                        have_original = bool(original_file) and original_file in st.session_state.original_files

                        template_src = meta.get('template_source_file')

                        export_choices = ["Current source only", "All sources (multiple sheets)"]
                        if have_original:
                            export_choices.insert(1, "📦 Original file, updated (same name, same sheets)")
                        if template_src:
                            export_choices.insert(1, "📐 Styled to match template (font/color/table)")

                        export_option = st.radio("Export:", export_choices, horizontal=True, key="export_option")

                        if export_option == "Current source only":
                            default_stem = Path(original_file).stem if original_file else source_name
                            output_filename = st.text_input(
                                "Output file name (without extension)", default_stem, key="export_filename_input",
                                help="Defaults to your original uploaded file's name so you can drop the "
                                     "download right back where the original was."
                            )
                            output_filename = output_filename.strip() or default_stem

                            col1, col2 = st.columns(2)
                            with col1:
                                csv = df.to_csv(index=False)
                                st.download_button(
                                    "📊 Download CSV", data=csv, file_name=f"{output_filename}.csv",
                                    mime="text/csv", use_container_width=True, key="download_csv_btn"
                                )
                            with col2:
                                excel_data = to_excel_bytes({source_name: df})
                                st.download_button(
                                    "📊 Download Excel", data=excel_data, file_name=f"{output_filename}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True, key="download_xlsx_btn"
                                )

                        elif export_option == "📐 Styled to match template (font/color/table)":
                            st.info(
                                f"This writes your built output into a copy of **'{template_src['name']}'**, "
                                "reusing its fonts, colors, borders, and (if present) its Excel Table styling "
                                "for every data cell — instead of a plain, unstyled sheet."
                            )
                            try:
                                styled_bytes = build_styled_template_excel(
                                    template_src['bytes'], template_src['sheet'], template_src['header_row'],
                                    list(df.columns), df
                                )
                                st.download_button(
                                    f"📐 Download Styled Excel (matches '{template_src['name']}')",
                                    data=styled_bytes,
                                    file_name=f"{Path(template_src['name']).stem}_filled.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True, key="download_styled_template_btn"
                                )
                            except Exception as e:
                                st.error(f"Couldn't build the styled version: {e}")

                        elif export_option == "📦 Original file, updated (same name, same sheets)":
                            st.info(
                                f"This rebuilds **'{original_file}'** exactly as it was — same file name, "
                                "same number of sheets, same sheet names, same order — but the sheet(s) "
                                "you've processed use your updated data. Any sheet from this file you never "
                                "touched is included exactly as it was originally uploaded, untouched."
                            )
                            file_bytes, ext_used = build_original_file_bytes(
                                original_file, st.session_state.source_metadata, st.session_state.data_sources
                            )
                            if file_bytes is None:
                                st.error(
                                    "The original file's raw bytes aren't available in this session anymore "
                                    "(e.g. the app restarted since you uploaded it). Re-upload the original "
                                    "file and re-run your recipe via Automation to use this export mode."
                                )
                            else:
                                mime = "text/csv" if ext_used == ".csv" else \
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                st.download_button(
                                    f"📦 Download '{original_file}' (updated)",
                                    data=file_bytes,
                                    file_name=f"{Path(original_file).stem}{ext_used}",
                                    mime=mime, use_container_width=True, key="download_original_btn"
                                )

                        else:
                            excel_data = to_excel_bytes(st.session_state.data_sources)
                            st.download_button(
                                "📊 Download All Sources",
                                data=excel_data,
                                file_name=f"all_sources_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, key="download_all_btn"
                            )

                        st.info(f"📊 Exporting {len(df)} rows and {len(df.columns)} columns")

                    st.markdown('</div>', unsafe_allow_html=True)
            else:
                # Welcome / Select data source prompt
                st.markdown(f"""
                <div class="welcome-box" style="text-align: center;">
                    <h2>📊 Select a Data Source</h2>
                    <p style="font-size: 1.1rem; color: #cbd5e1; margin-bottom: 15px;">Choose any source from the left panel to start processing</p>
                    <span style="background: rgba(0, 242, 254, 0.1); border: 1px solid rgba(0, 242, 254, 0.2); padding: 6px 16px; border-radius: 20px; font-size: 0.9rem; color: #00f2fe; font-weight: 500;">
                        {len(st.session_state.data_sources)} active data sources loaded
                    </span>
                </div>
                """, unsafe_allow_html=True)
                st.subheader("📋 All Loaded Sources")
                for name, df in st.session_state.data_sources.items():
                    with st.expander(f"📊 {name} ({len(df)} rows, {len(df.columns)} cols)"):
                        st.dataframe(df.head(), use_container_width=True)

elif page == "Automation":
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown('<h2 style="font-family: \'Outfit\', sans-serif; color: #00f2fe; margin-top: 0;">🤖 Run Automation Replay</h2>', unsafe_allow_html=True)
    st.markdown('<p style="font-size: 1rem; color: #cbd5e1; margin-bottom: 25px;">Upload a recipe JSON you downloaded earlier and execute it instantly against new data worksheets.</p>', unsafe_allow_html=True)
    
    recipe_file = st.file_uploader("Upload a recipe (.json)", type=['json'], key="recipe_uploader")
    if recipe_file is not None:
        try:
            st.session_state.loaded_recipe = json.load(recipe_file)
            st.session_state.recipe_mapping = {}
        except Exception as e:
            st.error(f"Couldn't read recipe file: {e}")
            st.session_state.loaded_recipe = None

    if st.session_state.loaded_recipe and st.session_state.data_sources:
        recipe = st.session_state.loaded_recipe
        needed_sources = extract_referenced_sources(recipe)
        st.write(f"This recipe needs **{len(needed_sources)}** source(s) and has **{len(recipe.get('operations', []))}** step(s).")

        loaded_names = list(st.session_state.data_sources.keys())
        for needed in needed_sources:
            default_index = loaded_names.index(needed) if needed in loaded_names else 0
            chosen = st.selectbox(f"Map '{needed}' to:", loaded_names, index=default_index, key=f"map_{needed}")
            st.session_state.recipe_mapping[needed] = chosen

        if st.button("▶️ Run Automation", use_container_width=True, key="run_automation_btn"):
            mapping = st.session_state.recipe_mapping
            primary_new = mapping.get(recipe.get('primary_source'))
            if not primary_new:
                st.error("Please map the primary source before running.")
            else:
                remapped_ops = remap_operations(recipe.get('operations', []), mapping)
                primary_df = st.session_state.data_sources[primary_new]
                result_df = execute_all_operations(primary_df, remapped_ops, st.session_state.data_sources)
                result_name = f"{primary_new}_automated_{datetime.now().strftime('%H%M%S')}"
                st.session_state.data_sources[result_name] = result_df
                st.session_state.source_snapshots[result_name] = primary_df.copy()
                st.session_state.source_metadata[result_name] = {
                    'file': 'Automation', 'sheet': 'Result',
                    'rows': len(result_df), 'columns': len(result_df.columns)
                }
                st.session_state.active_source = result_name
                st.session_state.operations = list(remapped_ops)
                st.session_state.operations_undo_stack = []
                st.success(f"✅ Automation complete! View the output in your Workspace: '{result_name}'")
                st.session_state.current_page = "Workspace"
                st.rerun()
    elif st.session_state.loaded_recipe and not st.session_state.data_sources:
        st.warning("⚠️ No data sources loaded. Please load the necessary reference files in the Workspace first.")
    else:
        st.info("💡 Load a recipe JSON to begin mapping.")
    st.markdown('</div>', unsafe_allow_html=True)

elif page == "Batch Automation":
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown('<h2 style="font-family: \'Outfit\', sans-serif; color: #9b51e0; margin-top: 0;">📦 Batch Automation Processor</h2>', unsafe_allow_html=True)
    st.markdown('<p style="font-size: 1rem; color: #cbd5e1; margin-bottom: 25px;">Apply a saved cleaning recipe on multiple files at once. Each file yields a distinct output sheet.</p>', unsafe_allow_html=True)
    
    batch_recipe_file = st.file_uploader("Upload a recipe (.json)", type=['json'], key="batch_recipe_uploader")
    batch_files = st.file_uploader("Upload files to process — one run of the recipe per file", type=['csv', 'xlsx', 'xls'], accept_multiple_files=True, key="batch_files_uploader")

    if batch_recipe_file is not None and batch_files:
        try:
            batch_recipe = json.load(batch_recipe_file)
        except Exception as e:
            st.error(f"Couldn't read recipe file: {e}")
            batch_recipe = None

        if batch_recipe:
            primary_name = batch_recipe.get('primary_source')
            needed_sources = extract_referenced_sources(batch_recipe)
            other_needed = [s for s in needed_sources if s != primary_name]

            st.write(f"Each uploaded file will be used as **'{primary_name}'**. This recipe has **{len(batch_recipe.get('operations', []))}** step(s).")

            ref_mapping = {}
            loaded_names = list(st.session_state.data_sources.keys())
            if other_needed:
                st.caption("This recipe also needs reference source(s) — map them to data you already have loaded:")
                if not loaded_names:
                    st.warning("Load the needed reference file(s) first in the Workspace.")
                for needed in other_needed:
                    default_index = loaded_names.index(needed) if needed in loaded_names else 0
                    if loaded_names:
                        ref_mapping[needed] = st.selectbox(f"Map '{needed}' to:", loaded_names, index=default_index, key=f"batch_map_{needed}")

            if st.button("▶️ Run Batch", use_container_width=True, key="run_batch_btn"):
                results, errors = {}, {}
                remapped_ops = remap_operations(batch_recipe.get('operations', []), ref_mapping)
                skipped_sheets = {}
                suspect_headers = []
                for f in batch_files:
                    try:
                        file_stem = Path(f.name).stem
                        # Check if file is password protected
                        ext = Path(f.name).suffix.lower()
                        file_bytes = f.getvalue()
                        
                        if ext in ('.xlsx', '.xls') and is_excel_password_protected(file_bytes):
                            # Try with cached password if available
                            password = st.session_state.password_cache.get(f.name)
                            if password:
                                sheets, status = load_excel_with_password(file_bytes, password)
                            else:
                                errors[f.name] = "File is password protected. Please unlock it in the Data Sources tab first."
                                continue
                        else:
                            sheets = MultiSheetProcessor.load_single_sheet(f)
                            status = "success" if sheets else "failed"
                        
                        if not sheets or status != "success":
                            errors[f.name] = "Could not read file"
                            continue
                        
                        sheet_names = list(sheets.keys())
                        primary_df = sheets[sheet_names[0]]
                        if len(sheet_names) > 1:
                            skipped_sheets[f.name] = sheet_names[1:]
                        if columns_look_unnamed_heavy(primary_df.columns):
                            suspect_headers.append(f.name)
                        result_df = execute_all_operations(primary_df, remapped_ops, st.session_state.data_sources)
                        results[f"{file_stem}_result"] = result_df
                    except Exception as e:
                        errors[f.name] = str(e)

                st.session_state.batch_results = results
                if results:
                    st.success(f"✅ Batch complete: {len(results)} of {len(batch_files)} file(s) processed.")
                if suspect_headers:
                    st.warning(
                        "⚠️ These file(s) have a lot of unnamed/blank columns after loading, which "
                        "usually means the real header row isn't the first row (a title or blank row "
                        "above it). Batch mode always uses row 0 as the header — load these files "
                        "individually first (top of sidebar) to confirm the correct header row, or "
                        "fix the file directly: " + ", ".join(suspect_headers)
                    )
                if skipped_sheets:
                    st.warning(
                        "⚠️ Only the first sheet of each file is used as input. Extra sheet(s) "
                        "were NOT processed: " +
                        "; ".join(f"{fname} → {', '.join(names)}" for fname, names in skipped_sheets.items())
                    )
                if errors:
                    st.error(f"⚠️ {len(errors)} file(s) failed:")
                    for fname, err in errors.items():
                        st.write(f"- **{fname}**: {err}")
                st.rerun()

    if st.session_state.batch_results:
        st.divider()
        st.subheader("📦 Batch Results")
        for name, rdf in st.session_state.batch_results.items():
            with st.expander(f"📊 {name} ({len(rdf)} rows, {len(rdf.columns)} cols)"):
                st.dataframe(rdf.head(20), use_container_width=True)

        combined_excel = to_excel_bytes(st.session_state.batch_results)
        st.download_button(
            "📥 Download All Batch Results (one Excel workbook)",
            data=combined_excel,
            file_name=f"batch_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="download_batch_all_btn"
        )

        if st.button("➕ Load batch results as data sources", use_container_width=True, key="load_batch_as_sources_btn"):
            for name, rdf in st.session_state.batch_results.items():
                st.session_state.data_sources[name] = rdf
                st.session_state.source_snapshots[name] = rdf.copy()
                st.session_state.source_metadata[name] = {
                    'file': 'Batch Automation', 'sheet': 'Result',
                    'rows': len(rdf), 'columns': len(rdf.columns)
                }
            st.success("Loaded batch results into your data sources.")
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

elif page == "Manual":
    features_html = """
    <!DOCTYPE html>
    <html>
    <head>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap');
        body {
            margin: 0;
            padding: 0;
            background: transparent;
            font-family: 'Inter', sans-serif;
            color: #cbd5e1;
            overflow: hidden;
        }
        .carousel-card {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.03) 0%, rgba(255, 255, 255, 0.005) 100%);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid rgba(255, 255, 255, 0.05);
            border-top: 1px solid rgba(0, 242, 254, 0.20);
            border-left: 1px solid rgba(0, 242, 254, 0.10);
            border-radius: 12px;
            padding: 30px;
            height: 250px;
            box-sizing: border-box;
            position: relative;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.4);
        }
        .slide {
            display: none;
            animation: fade 0.6s ease-in-out;
            width: 100%;
        }
        .slide.active {
            display: block;
        }
        @keyframes fade {
            from { opacity: 0; transform: scale(0.98); }
            to { opacity: 1; transform: scale(1); }
        }
        .title {
            font-family: 'Outfit', sans-serif;
            color: #00f2fe;
            font-size: 1.35rem;
            font-weight: 700;
            margin: 0 0 12px 0;
            text-shadow: 0 0 10px rgba(0, 242, 254, 0.2);
        }
        .desc {
            font-size: 0.95rem;
            color: #cbd5e1;
            line-height: 1.6;
            margin: 0 auto;
            max-width: 600px;
        }
        .nav-btn {
            position: absolute;
            top: 50%;
            transform: translateY(-50%);
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid rgba(255, 255, 255, 0.06);
            color: #8a99ad;
            font-size: 18px;
            width: 40px;
            height: 40px;
            border-radius: 50%;
            cursor: pointer;
            transition: all 0.2s ease;
            user-select: none;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .nav-btn:hover {
            background: linear-gradient(135deg, rgba(0, 242, 254, 0.15) 0%, rgba(155, 81, 224, 0.15) 100%);
            border-color: rgba(0, 242, 254, 0.4);
            color: #ffffff;
            box-shadow: 0 0 15px rgba(0, 242, 254, 0.3);
        }
        .prev { left: 15px; }
        .next { right: 15px; }
        .dots-container {
            position: absolute;
            bottom: 15px;
            display: flex;
            gap: 8px;
        }
        .dot {
            width: 8px;
            height: 8px;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.12);
            cursor: pointer;
            transition: all 0.2s ease;
        }
        .dot.active {
            background: #00f2fe;
            box-shadow: 0 0 8px #00f2fe;
        }
    </style>
    </head>
    <body>
    <div class="carousel-card">
        <div class="slide active">
            <div class="title">📚 Multi-Sheet Loader</div>
            <div class="desc">Load multiple Excel sheets and standalone CSV files automatically. Excel sheets are imported as separate files so you can run cross-sheet processes easily.</div>
        </div>
        <div class="slide">
            <div class="title">🔗 Relationships Manager</div>
            <div class="desc">Connect files dynamically using standard SQL joins (Inner, Left, Right, Outer) on shared key columns to create merged master data sources.</div>
        </div>
        <div class="slide">
            <div class="title">🔄 Cross-Sheet Formula Engine</div>
            <div class="desc">Apply advanced spreadsheet-style formulas, including native XLOOKUP and VLOOKUP functions, to map and lookup data across loaded sheets.</div>
        </div>
        <div class="slide">
            <div class="title">📊 Visual Recipe Timeline</div>
            <div class="desc">Record all cleaning and filter operations into a visual sequential recipe. Easily reorder, edit, toggle, or delete steps in real time.</div>
        </div>
        <div class="slide">
            <div class="title">🤖 Parallel Automation</div>
            <div class="desc">Save your recorded recipe and replay it instantly on brand-new files, eliminating the need to rebuild data cleaning processes.</div>
        </div>
        <div class="slide">
            <div class="title">🧩 Template Builder</div>
            <div class="desc">Configure exact column header output templates. Map columns via direct copy, lookup, custom formulas, or static values.</div>
        </div>

        <div class="nav-btn prev" onclick="moveSlide(-1)">&#10094;</div>
        <div class="nav-btn next" onclick="moveSlide(1)">&#10095;</div>

        <div class="dots-container">
            <div class="dot active" onclick="setSlide(0)"></div>
            <div class="dot" onclick="setSlide(1)"></div>
            <div class="dot" onclick="setSlide(2)"></div>
            <div class="dot" onclick="setSlide(3)"></div>
            <div class="dot" onclick="setSlide(4)"></div>
            <div class="dot" onclick="setSlide(5)"></div>
        </div>
    </div>
    <script>
        let currentIdx = 0;
        const slides = document.querySelectorAll('.slide');
        const dots = document.querySelectorAll('.dot');
        let timer = setInterval(autoPlay, 5000);

        function updateCarousel() {
            slides.forEach((slide, idx) => {
                slide.classList.toggle('active', idx === currentIdx);
            });
            dots.forEach((dot, idx) => {
                dot.classList.toggle('active', idx === currentIdx);
            });
        }

        function moveSlide(dir) {
            clearInterval(timer);
            currentIdx = (currentIdx + dir + slides.length) % slides.length;
            updateCarousel();
            timer = setInterval(autoPlay, 5000);
        }

        function setSlide(idx) {
            clearInterval(timer);
            currentIdx = idx;
            updateCarousel();
            timer = setInterval(autoPlay, 5000);
        }

        function autoPlay() {
            currentIdx = (currentIdx + 1) % slides.length;
            updateCarousel();
        }
    </script>
    </body>
    </html>
    """

    steps_html = """
    <!DOCTYPE html>
    <html>
    <head>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=Inter:wght@300;400;500;600;700&display=swap');
        body {
            margin: 0;
            padding: 0;
            background: transparent;
            font-family: 'Inter', sans-serif;
            color: #cbd5e1;
            overflow: hidden;
        }
        .carousel-card {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.03) 0%, rgba(255, 255, 255, 0.005) 100%);
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid rgba(255, 255, 255, 0.05);
            border-top: 1px solid rgba(155, 81, 224, 0.20);
            border-left: 1px solid rgba(155, 81, 224, 0.10);
            border-radius: 12px;
            padding: 30px;
            height: 250px;
            box-sizing: border-box;
            position: relative;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.4);
        }
        .slide {
            display: none;
            animation: fade 0.6s ease-in-out;
            width: 100%;
        }
        .slide.active {
            display: block;
        }
        @keyframes fade {
            from { opacity: 0; transform: scale(0.98); }
            to { opacity: 1; transform: scale(1); }
        }
        .badge {
            background: rgba(155, 81, 224, 0.15) !important;
            border: 1px solid rgba(155, 81, 224, 0.3) !important;
            color: #d6b3ff !important;
            padding: 4px 10px !important;
            border-radius: 6px !important;
            font-size: 0.75rem;
            font-weight: 600;
            display: inline-block;
            margin-bottom: 12px;
        }
        .title {
            font-family: 'Outfit', sans-serif;
            color: #9b51e0;
            font-size: 1.35rem;
            font-weight: 700;
            margin: 0 0 10px 0;
            text-shadow: 0 0 10px rgba(155, 81, 224, 0.2);
        }
        .desc {
            font-size: 0.95rem;
            color: #cbd5e1;
            line-height: 1.6;
            margin: 0 auto;
            max-width: 600px;
        }
        .nav-btn {
            position: absolute;
            top: 50%;
            transform: translateY(-50%);
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid rgba(255, 255, 255, 0.06);
            color: #8a99ad;
            font-size: 18px;
            width: 40px;
            height: 40px;
            border-radius: 50%;
            cursor: pointer;
            transition: all 0.2s ease;
            user-select: none;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .nav-btn:hover {
            background: linear-gradient(135deg, rgba(155, 81, 224, 0.15) 0%, rgba(0, 242, 254, 0.15) 100%);
            border-color: rgba(155, 81, 224, 0.4);
            color: #ffffff;
            box-shadow: 0 0 15px rgba(155, 81, 224, 0.3);
        }
        .prev { left: 15px; }
        .next { right: 15px; }
        .dots-container {
            position: absolute;
            bottom: 15px;
            display: flex;
            gap: 8px;
        }
        .dot {
            width: 8px;
            height: 8px;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.12);
            cursor: pointer;
            transition: all 0.2s ease;
        }
        .dot.active {
            background: #9b51e0;
            box-shadow: 0 0 8px #9b51e0;
        }
    </style>
    </head>
    <body>
    <div class="carousel-card">
        <div class="slide active">
            <div class="badge">Step 1</div>
            <div class="title">Load Data</div>
            <div class="desc">Upload your CSV or Excel files in the Data Sources page and click <strong>🔄 Load All Files</strong>. Every sheet will register as a distinct data source.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 2</div>
            <div class="title">Build Recipe</div>
            <div class="desc">Select an active sheet and use the <strong>⚡ Data Operations</strong> panel to append filters, sort data, group columns, or rename fields.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 3</div>
            <div class="title">Filter Rows</div>
            <div class="desc">Apply single or multi-condition filters using AND/OR logic. Or filter values based on reference columns from other sheets.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 4</div>
            <div class="title">Add Formulas</div>
            <div class="desc">Reference columns with square brackets, e.g. <code>[Price] * [Qty]</code>. Support includes <code>IF/THEN/ELSE</code> and custom lookup calls.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 5</div>
            <div class="title">Join Sheets</div>
            <div class="desc">Establish relationships between different tables. Create custom VLOOKUPs or merge full tables through the Join manager.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 6</div>
            <div class="title">Map Outputs</div>
            <div class="desc">Create output templates matching your reporting requirements, mapping cells from source sheets to target columns.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 7</div>
            <div class="title">Automate Reports</div>
            <div class="desc">Download your cleaning recipe as a JSON file, load it under <strong>🤖 Automation</strong>, and run it on new reporting cycles.</div>
        </div>
        <div class="slide">
            <div class="badge">Step 8</div>
            <div class="title">Export Workbooks</div>
            <div class="desc">Download your processed worksheets as single-sheet CSV/Excel files, or package all data sources into a multi-tab workbook.</div>
        </div>

        <div class="nav-btn prev" onclick="moveSlide(-1)">&#10094;</div>
        <div class="nav-btn next" onclick="moveSlide(1)">&#10095;</div>

        <div class="dots-container">
            <div class="dot active" onclick="setSlide(0)"></div>
            <div class="dot" onclick="setSlide(1)"></div>
            <div class="dot" onclick="setSlide(2)"></div>
            <div class="dot" onclick="setSlide(3)"></div>
            <div class="dot" onclick="setSlide(4)"></div>
            <div class="dot" onclick="setSlide(5)"></div>
            <div class="dot" onclick="setSlide(6)"></div>
            <div class="dot" onclick="setSlide(7)"></div>
        </div>
    </div>
    <script>
        let currentIdx = 0;
        const slides = document.querySelectorAll('.slide');
        const dots = document.querySelectorAll('.dot');
        let timer = setInterval(autoPlay, 5000);

        function updateCarousel() {
            slides.forEach((slide, idx) => {
                slide.classList.toggle('active', idx === currentIdx);
            });
            dots.forEach((dot, idx) => {
                dot.classList.toggle('active', idx === currentIdx);
            });
        }

        function moveSlide(dir) {
            clearInterval(timer);
            currentIdx = (currentIdx + dir + slides.length) % slides.length;
            updateCarousel();
            timer = setInterval(autoPlay, 5000);
        }

        function setSlide(idx) {
            clearInterval(timer);
            currentIdx = idx;
            updateCarousel();
            timer = setInterval(autoPlay, 5000);
        }

        function autoPlay() {
            currentIdx = (currentIdx + 1) % slides.length;
            updateCarousel();
        }
    </script>
    </body>
    </html>
    """

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<h3 style="font-family: \'Outfit\', sans-serif; color: #00f2fe; text-align: center; margin-top: 10px; margin-bottom: 20px;">✨ Builder Core Features</h3>', unsafe_allow_html=True)
        st.components.v1.html(features_html, height=270)
    with col2:
        st.markdown('<h3 style="font-family: \'Outfit\', sans-serif; color: #9b51e0; text-align: center; margin-top: 10px; margin-bottom: 20px;">📖 Step-by-Step User Guide</h3>', unsafe_allow_html=True)
        st.components.v1.html(steps_html, height=270)

st.divider()
st.markdown("""
<div style="text-align: center; color: #64748b; font-size: 0.85rem; padding: 15px 0;">
    DataWiz v2.1 | Powered by Streamlit, Pandas & Advanced UI Engine
</div>
""", unsafe_allow_html=True)